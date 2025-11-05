"""
Data extraction functions
File ini berisi semua fungsi untuk mengekstrak data dari halaman web
"""
import time
import re
from datetime import datetime
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

def get_stock_value_direct(driver):
    """
    ============================================
    FUNGSI AMBIL DATA STOK DARI DASHBOARD
    ============================================
    
    Fungsi ini mengambil data stok dari dashboard merchant dengan menggunakan
    direct class selector yang sudah dioptimasi untuk performa maksimal.
    
    Proses yang dilakukan:
    1. Mencari elemen stok menggunakan class selector yang sudah diketahui
    2. Extract text dari elemen stok
    3. Menggunakan regex untuk mengambil angka pertama dari text
    4. Return nilai stok dalam format string
    
    Fungsi ini sangat cepat karena menggunakan direct selector tanpa iterasi
    dan sudah terbukti berhasil berdasarkan debugging sebelumnya.
    
    Args:
        driver: WebDriver object yang sudah login dan berada di dashboard
    
    Returns:
        str: Nilai stok dalam format string (contoh: "150"), atau None jika gagal
    """
    print("\n📦 === AMBIL DATA STOK LANGSUNG ===")
    try:
        time.sleep(1.5)
        print("🚀 Mengambil stok langsung menggunakan class yang sudah diketahui...")
        
        # Langsung ambil dengan class yang sudah terbukti berhasil
        element = driver.find_element(By.CLASS_NAME, "styles_summaryProductCard__Uv3IK")
        text = element.text.strip()
        
        print(f"✅ Elemen stok ditemukan langsung!")
        print(f"📝 Text: '{text}'")
        
        # Langsung extract angka pertama tanpa validasi tambahan
        numbers = re.findall(r'\d+', text)
        if numbers:
            stock_value = numbers[0]
            print(f"🔢 Angka: {numbers}")
            print(f"📊 Nilai Stok: {stock_value}")
            return stock_value
        else:
            print("❌ Tidak ada angka ditemukan dalam text")
            return None
            
    except Exception as e:
        print(f"❌ Error mengambil data stok: {str(e)}")
        return None

def get_tabung_terjual_direct(driver):
    """
    ============================================
    FUNGSI AMBIL DATA TABUNG TERJUAL DARI LAPORAN PENJUALAN
    ============================================
    
    Fungsi ini mengambil data tabung terjual dari halaman Laporan Penjualan dengan
    menggunakan direct selector yang sudah dioptimasi untuk performa maksimal.
    
    Proses yang dilakukan:
    1. Mencari elemen dengan class "mantine-Text-root" yang berisi data tabung
    2. Filter elemen yang mengandung kata "tabung" dan angka
    3. Skip elemen yang merupakan header "Total Tabung LPG 3 Kg Terjual"
    4. Extract angka dari text yang ditemukan
    5. Return data dalam format "X Tabung"
    6. Fallback ke XPath selector jika class selector gagal
    
    Fungsi ini sangat cepat karena menggunakan direct selector tanpa iterasi
    dan sudah terbukti berhasil berdasarkan debugging sebelumnya.
    
    Args:
        driver: WebDriver object yang sudah berada di halaman Laporan Penjualan
    
    Returns:
        str: Data tabung terjual dalam format "X Tabung" (contoh: "28 Tabung"), 
             atau None jika gagal mengambil data
    """
    print("\n📊 === AMBIL DATA TABUNG TERJUAL LANGSUNG ===")
    
    try:
        time.sleep(1.5)
        print("🚀 Mengambil data tabung terjual langsung menggunakan lokasi yang sudah diketahui...")
        
        try:
            # Coba dengan class yang sudah diketahui
            elements = driver.find_elements(By.CLASS_NAME, "mantine-Text-root")
            
            for element in elements:
                text = element.text.strip()
                if text and 'tabung' in text.lower() and any(char.isdigit() for char in text):
                    # Skip jika ini adalah text "Total Tabung LPG 3 Kg Terjual"
                    if 'total' in text.lower() and 'terjual' in text.lower():
                        continue
                    
                    print(f"✅ Data tabung terjual ditemukan langsung!")
                    print(f"📝 Text: '{text}'")
                    
                    # Ekstrak angka langsung
                    numbers = re.findall(r'\d+', text)
                    if numbers:
                        tabung_value = numbers[0]
                        clean_text = f"{tabung_value} Tabung"
                        print(f"🔢 Angka: {numbers}")
                        print(f"📊 Jumlah Tabung Terjual: {tabung_value}")
                        print(f"📝 Text Bersih: '{clean_text}'")
                        return clean_text
            
            print("❌ Data tabung terjual tidak ditemukan dengan class selector")
            return None
                
        except Exception as e:
            print(f"⚠️ Error dengan class selector: {str(e)}")
            
            # Fallback: gunakan XPath yang lebih spesifik berdasarkan text pattern
            print("🔄 Mencoba dengan XPath fallback yang lebih efisien...")
            try:
                fallback_selectors = [
                    "//*[contains(text(), 'Tabung')]",
                    "//div[contains(@class, 'text')]//*[contains(text(), 'tabung')]",
                    "//span[contains(text(), 'tabung')]",
                    "//p[contains(text(), 'tabung')]"
                ]
                for selector in fallback_selectors:
                    try:
                        elements = driver.find_elements(By.XPATH, selector)
                        for element in elements:
                            text = element.text.strip()
                            if (text and len(text) < 50 and 
                                'tabung' in text.lower() and 
                                any(char.isdigit() for char in text) and 
                                'total' not in text.lower()):
                                
                                print(f"✅ Data tabung terjual ditemukan dengan XPath fallback!")
                                print(f"📝 Text: '{text}'")
                                
                                numbers = re.findall(r'\d+', text)
                                if numbers:
                                    tabung_value = numbers[0]
                                    clean_text = f"{tabung_value} Tabung"
                                    print(f"🔢 Angka: {numbers}")
                                    print(f"📊 Jumlah Tabung Terjual: {tabung_value}")
                                    print(f"📝 Text Bersih: '{clean_text}'")
                                    return clean_text
                    except:
                        continue
                                
            except Exception as e2:
                print(f"❌ Error dengan XPath fallback: {str(e2)}")
                return None
            
    except Exception as e:
        print(f"❌ Error mengambil data tabung terjual: {str(e)}")
        return None

def get_customer_list_direct(driver):
    """Ambil data list pembeli langsung menggunakan lokasi yang sudah diketahui - lebih cepat"""
    print("\n👥 === AMBIL DATA LIST PEMBELI LANGSUNG ===")
    
    try:
        time.sleep(1.0)
        print("🚀 Mengambil data list pembeli langsung menggunakan lokasi yang sudah diketahui...")
        
        try:
            elements = None
            customer_selectors = [
                (By.CLASS_NAME, "styles_listTransactionRoot__pvz4r"),
                (By.CLASS_NAME, "mantine-1uguyhf"),
                (By.XPATH, "//div[@class='styles_listTransactionRoot__pvz4r mantine-1uguyhf']")
            ]
            for selector_type, selector_value in customer_selectors:
                try:
                    if selector_type == By.CLASS_NAME:
                        elements = driver.find_elements(selector_type, selector_value)
                        if elements:
                            break
                    elif selector_type == By.XPATH:
                        elements = driver.find_elements(selector_type, selector_value)
                        if elements:
                            break
                except:
                    continue
            
            if elements:
                print(f"✅ Ditemukan {len(elements)} container pembeli langsung!")
                
                customer_data = []
                for idx, element in enumerate(elements):
                    try:
                        text = element.text.strip()
                        if text and ('tabung' in text.lower() and 'lpg' in text.lower()):
                            print(f"✅ Container pembeli {idx+1} ditemukan langsung!")
                            print(f"📝 Data: '{text[:50]}...'")
                            
                            customer_data.append({
                                'text': text,
                                'element': element,
                                'element_info': {
                                    'tag_name': element.tag_name,
                                    'id': element.get_attribute('id'),
                                    'class': element.get_attribute('class'),
                                    'index': idx
                                }
                            })
                    except Exception as e:
                        print(f"⚠️ Error membaca container {idx+1}: {str(e)}")
                        continue
                
                if customer_data:
                    print(f"\n✅ DATA LIST PEMBELI BERHASIL DIAMBIL LANGSUNG:")
                    print(f"   📊 Total pembeli: {len(customer_data)}")
                    
                    # Tampilkan preview beberapa pembeli pertama
                    for idx, customer in enumerate(customer_data[:3]):
                        print(f"\n👥 === PEMBELI {idx+1} (Preview) ===")
                        print(f"   📝 Data: '{customer['text'][:100]}...'")
                        print(f"   🏷️ Tag: {customer['element_info']['tag_name']}")
                        print(f"   🎨 Class: '{customer['element_info']['class'] or 'N/A'}'")
                    
                    if len(customer_data) > 3:
                        print(f"\n... dan {len(customer_data) - 3} pembeli lainnya")
                    
                    return customer_data
                else:
                    print("❌ Tidak ada data pembeli yang valid ditemukan")
                    return None
            else:
                print("❌ Container pembeli tidak ditemukan dengan class yang diketahui")
                return None
                
        except Exception as e:
            print(f"❌ Error mengambil data list pembeli: {str(e)}")
            return None
            
    except Exception as e:
        print(f"❌ Error mengambil data list pembeli: {str(e)}")
        return None

def should_click_rekap_penjualan(tabung_terjual_text):
    """Decision function: apakah perlu klik Rekap Penjualan berdasarkan data tabung terjual"""
    print(f"\n🤔 === DECISION: APAKAH PERLU KLIK REKAP PENJUALAN? ===")
    
    try:
        if not tabung_terjual_text:
            print("❌ Data tabung terjual kosong - SKIP Rekap Penjualan")
            return False
        
        print(f"📊 Data tabung terjual: '{tabung_terjual_text}'")
        
        # Ekstrak angka dari text "28 Tabung" atau "0 Tabung"
        numbers = re.findall(r'\d+', tabung_terjual_text)
        
        if not numbers:
            print("❌ Tidak ada angka ditemukan - SKIP Rekap Penjualan")
            return False
        
        tabung_value = int(numbers[0])
        print(f"🔢 Nilai tabung terjual: {tabung_value}")
        
        if tabung_value == 0:
            print("❌ Tabung terjual = 0 - SKIP Rekap Penjualan")
            print("💡 Tidak ada penjualan, tidak perlu melihat rekap")
            return False
        elif tabung_value > 0:
            print(f"✅ Tabung terjual = {tabung_value} (> 0) - KLIK Rekap Penjualan")
            print(f"💡 Ada penjualan {tabung_value} tabung, perlu melihat rekap detail")
            return True
        else:
            print(f"⚠️ Nilai tabung tidak valid: {tabung_value} - SKIP Rekap Penjualan")
            return False
        
    except Exception as e:
        print(f"❌ Error dalam decision function: {str(e)}")
        print("⚠️ Error - SKIP Rekap Penjualan")
        return False


def read_nik_from_excel():
    """
    ============================================
    FUNGSI BACA NIK DARI FILE EXCEL
    ============================================
    
    Fungsi ini membaca daftar NIK dari file akun/NIK.xlsx
    dan mengembalikan list NIK yang siap digunakan.
    
    Returns:
        list: List NIK dalam format string, atau None jika gagal
    """
    print("\n📋 === BACA NIK DARI EXCEL ===")
    
    try:
        import openpyxl
        import os
        
        nik_file_path = "akun/NIK.xlsx"
        
        if not os.path.exists(nik_file_path):
            print(f"❌ File NIK tidak ditemukan: {nik_file_path}")
            return None
        
        print(f"📂 Membaca file NIK: {nik_file_path}")
        
        wb = openpyxl.load_workbook(nik_file_path)
        ws = wb.active
        
        print(f"📊 Sheet: {ws.title}")
        print(f"📊 Total rows: {ws.max_row}")
        
        nik_list = []
        
        # Baca NIK dari kolom A (mulai dari row 2, karena row 1 adalah header)
        for row in range(2, ws.max_row + 1):
            nik_value = ws.cell(row=row, column=1).value
            if nik_value and str(nik_value).strip():
                nik_list.append(str(nik_value).strip())
        
        print(f"✅ Berhasil membaca {len(nik_list)} NIK dari Excel")
        
        if nik_list:
            print(f"📝 NIK pertama: {nik_list[0]}")
            print(f"📝 NIK terakhir: {nik_list[-1]}")
        
        return nik_list
        
    except Exception as e:
        print(f"❌ Error membaca file NIK: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


def fill_nik_form_and_continue(driver, nik_list, start_index=0):
    """
    ============================================
    FUNGSI ISI FORM NIK DAN LANJUTKAN PENJUALAN
    ============================================
    
    Fungsi ini mengisi form NIK di halaman Catat Penjualan
    dan mengklik tombol "LANJUTKAN PENJUALAN".
    
    Args:
        driver: WebDriver object yang sudah berada di halaman Catat Penjualan
        nik_list (list): List NIK yang akan diisi
        start_index (int): Index NIK yang akan dimulai (default: 0)
    
    Returns:
        bool: True jika berhasil mengisi form dan klik lanjutkan, False jika gagal
    """
    print(f"\n📝 === ISI FORM NIK DAN LANJUTKAN PENJUALAN ===")
    
    try:
        if not nik_list or start_index >= len(nik_list):
            print("❌ Tidak ada NIK yang tersedia atau index melebihi batas")
            return False
        
        nik_to_use = nik_list[start_index]
        print(f"🔢 Menggunakan NIK index {start_index}: {nik_to_use}")
        
        # Tunggu popup muncul dan cari input field NIK dengan explicit wait
        print("⏳ Menunggu popup form NIK muncul...")
        nik_input = None
        try:
            # Gunakan explicit wait untuk menunggu elemen muncul (maksimal 10 detik)
            wait = WebDriverWait(driver, 10)
            nik_input = wait.until(
                EC.presence_of_element_located((By.XPATH, "//input[contains(@placeholder, 'NIK')]"))
            )
            print(f"✅ Input field NIK ditemukan dengan placeholder")
        except Exception as e:
            print(f"❌ Input field NIK tidak ditemukan setelah menunggu: {str(e)}")
            return False
        
        # Debug info (hanya jika berhasil)
        if nik_input:
            try:
                xpath = driver.execute_script("""
                    function absoluteXPath(el){
                      if(el.id) return '//*[@id="'+el.id+'"]';
                      const parts=[]; while(el && el.nodeType===1){
                        let ix=0, sib=el.previousSibling; while(sib){ if(sib.nodeType===1 && sib.nodeName===el.nodeName) ix++; sib=sib.previousSibling; }
                        parts.unshift(el.nodeName.toLowerCase()+'['+(ix+1)+']'); el=el.parentNode; }
                      return '//'+parts.join('/'); }
                    return absoluteXPath(arguments[0]);
                """, nik_input)
                css = driver.execute_script("""
                    function cssPath(el){ if (!(el instanceof Element)) return; const path=[]; while (el.nodeType===1){ let selector=el.nodeName.toLowerCase(); if (el.id){ selector+='#'+el.id; path.unshift(selector); break; } else { let sib=el, nth=1; while (sib=sib.previousElementSibling){ if (sib.nodeName.toLowerCase()==selector) nth++; } selector += ':nth-of-type('+nth+')'; path.unshift(selector); el=el.parentNode; } } return path.join(' > '); }
                    return cssPath(arguments[0]);
                """, nik_input)
                print(f"🔗 NIK Input XPath: {xpath}")
                print(f"🔗 NIK Input CSS: {css}")
            except Exception:
                pass
        
        if not nik_input:
            print("❌ Input field NIK tidak ditemukan")
            return False
        
        # Clear dan isi NIK
        print(f"📝 Mengisi NIK: {nik_to_use}")
        nik_input.clear()
        nik_input.send_keys(nik_to_use)
        
        # Tunggu sebentar untuk memastikan input terisi
        time.sleep(0.5)
        
        # Cari tombol "LANJUTKAN PENJUALAN" dengan direct approach
        continue_button = None
        try:
            # Coba cari tombol dengan text "LANJUTKAN PENJUALAN" langsung
            buttons = driver.find_elements(By.XPATH, "//button[contains(text(), 'LANJUTKAN PENJUALAN')]")
            if buttons:
                continue_button = buttons[0]  # Ambil yang pertama
                print(f"✅ Tombol LANJUTKAN PENJUALAN ditemukan: '{continue_button.text}'")
            else:
                # Fallback ke text "LANJUTKAN" saja
                buttons = driver.find_elements(By.XPATH, "//button[contains(text(), 'LANJUTKAN')]")
                if buttons:
                    continue_button = buttons[0]
                    print(f"✅ Tombol LANJUTKAN ditemukan: '{continue_button.text}'")
                else:
                    # Fallback ke class search
                    buttons = driver.find_elements(By.CLASS_NAME, "mantine-Button-root")
                    for button in buttons:
                        if 'LANJUTKAN' in button.text.strip().upper():
                            continue_button = button
                            print(f"✅ Tombol LANJUTKAN ditemukan dengan class: '{button.text}'")
                            break
        except Exception as e:
            print(f"⚠️ Error mencari tombol LANJUTKAN: {str(e)}")
        
        # Debug info (hanya jika berhasil)
        if continue_button:
            try:
                xpath = driver.execute_script("""
                    function absoluteXPath(el){
                      if(el.id) return '//*[@id="'+el.id+'"]';
                      const parts=[]; while(el && el.nodeType===1){
                        let ix=0, sib=el.previousSibling; while(sib){ if(sib.nodeType===1 && sib.nodeName===el.nodeName) ix++; sib=sib.previousSibling; }
                        parts.unshift(el.nodeName.toLowerCase()+'['+(ix+1)+']'); el=el.parentNode; }
                      return '//'+parts.join('/'); }
                    return absoluteXPath(arguments[0]);
                """, continue_button)
                css = driver.execute_script("""
                    function cssPath(el){ if (!(el instanceof Element)) return; const path=[]; while (el.nodeType===1){ let selector=el.nodeName.toLowerCase(); if (el.id){ selector+='#'+el.id; path.unshift(selector); break; } else { let sib=el, nth=1; while (sib=sib.previousElementSibling){ if (sib.nodeName.toLowerCase()==selector) nth++; } selector += ':nth-of-type('+nth+')'; path.unshift(selector); el=el.parentNode; } } return path.join(' > '); }
                    return cssPath(arguments[0]);
                """, continue_button)
                print(f"🔗 CEK LANJUTKAN XPath: {xpath}")
                print(f"🔗 CEK LANJUTKAN CSS: {css}")
            except Exception:
                pass
        
        if not continue_button:
            print("❌ Tombol LANJUTKAN PENJUALAN tidak ditemukan")
            return False
        
        # Klik tombol LANJUTKAN PENJUALAN
        print("🚀 Mengklik tombol LANJUTKAN PENJUALAN...")
        try:
            # Coba klik normal dulu
            continue_button.click()
        except Exception as click_error:
            print(f"⚠️ Klik normal gagal: {str(click_error)}")
            print("🔄 Mencoba JavaScript click...")
            try:
                # Fallback: JavaScript click
                driver.execute_script("arguments[0].click();", continue_button)
                print("✅ JavaScript click berhasil!")
            except Exception as js_error:
                print(f"❌ JavaScript click juga gagal: {str(js_error)}")
                raise js_error
        
        print("✅ Berhasil mengisi NIK dan mengklik LANJUTKAN PENJUALAN!")
        time.sleep(0.3)  # Tunggu halaman load (reduced untuk total delay maksimal 1 detik)

        # Coba langsung klik CEK PESANAN terlebih dahulu (optimasi: jika berhasil langsung lanjut)
        import time as time_module
        cek_start_time = time_module.time()
        print("🔍 Mencoba klik CEK PESANAN langsung...")
        try:
            # Panggil fungsi click_cek_pesanan yang sudah ada di file yang sama
            cek_pesanan_success = click_cek_pesanan(driver)
            cek_elapsed = time_module.time() - cek_start_time
            if cek_pesanan_success:
                print(f"✅ CEK PESANAN berhasil diklik langsung - lanjutkan proses (waktu: {cek_elapsed:.2f} detik)")
                return True
            else:
                print(f"⚠️ CEK PESANAN tidak ditemukan atau gagal - langsung cek TUTUP (waktu: {cek_elapsed:.2f} detik)")
                # Langsung skip ke TUTUP tanpa delay
        except Exception as e:
            cek_elapsed = time_module.time() - cek_start_time
            print(f"⚠️ Error saat mencoba klik CEK PESANAN: {str(e)} - langsung cek TUTUP (waktu: {cek_elapsed:.2f} detik)")
            # Langsung skip ke TUTUP tanpa delay
        
        # Jika CEK PESANAN gagal, langsung cek TUTUP tanpa delay
        try:
            # Optimasi: Cari tombol TUTUP dengan selector yang lebih cepat dan spesifik
            target_tutup = None
            tutup_start_time = time_module.time()
            max_tutup_time = 0.5  # Maksimal 0.5 detik untuk mencari TUTUP
            
            # Prioritaskan selector yang lebih cepat (handle whitespace, nested text, dan case-insensitive)
            fast_selectors = [
                # Exact match dengan normalize-space (case-insensitive)
                (By.XPATH, "//button[translate(normalize-space(text()), 'abcdefghijklmnopqrstuvwxyz', 'ABCDEFGHIJKLMNOPQRSTUVWXYZ')='TUTUP']"),  # Exact match normalized (prioritas 1)
                (By.XPATH, "//button[normalize-space(text())='TUTUP' or normalize-space(text())='Tutup' or normalize-space(text())='tutup']"),  # Variasi case (prioritas 2)
                # Contains dengan normalize-space
                (By.XPATH, "//button[contains(translate(normalize-space(.), 'abcdefghijklmnopqrstuvwxyz', 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'), 'TUTUP')]"),  # Contains normalized (prioritas 3)
                # Link dengan variasi yang sama
                (By.XPATH, "//a[translate(normalize-space(text()), 'abcdefghijklmnopqrstuvwxyz', 'ABCDEFGHIJKLMNOPQRSTUVWXYZ')='TUTUP']"),  # Link exact match (prioritas 4)
                (By.XPATH, "//a[normalize-space(text())='TUTUP' or normalize-space(text())='Tutup' or normalize-space(text())='tutup']"),  # Link variasi case (prioritas 5)
                (By.XPATH, "//a[contains(translate(normalize-space(.), 'abcdefghijklmnopqrstuvwxyz', 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'), 'TUTUP')]"),  # Link contains (prioritas 6)
                # Role button
                (By.XPATH, "//*[@role='button' and contains(translate(normalize-space(.), 'abcdefghijklmnopqrstuvwxyz', 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'), 'TUTUP')]"),  # Role button (prioritas 7)
                # Fallback: cari semua button/link yang mengandung TUTUP (case-insensitive, normalize-space)
                (By.XPATH, "//*[self::button or self::a or @role='button'][contains(translate(normalize-space(.), 'abcdefghijklmnopqrstuvwxyz', 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'), 'TUTUP')]"),  # Semua button/link (prioritas 8)
            ]
            
            print(f"🔍 Mencari tombol TUTUP dengan {len(fast_selectors)} selector...")
            
            # Coba selector cepat terlebih dahulu
            for idx, (selector_type, selector_value) in enumerate(fast_selectors, 1):
                # Cek timeout sebelum setiap selector
                elapsed = time_module.time() - tutup_start_time
                if elapsed >= max_tutup_time:
                    print(f"  ⏱️ Timeout tercapai saat selector #{idx}")
                    break
                    
                try:
                    selector_attempt_start = time_module.time()
                    elements = driver.find_elements(selector_type, selector_value)
                    selector_attempt_elapsed = time_module.time() - selector_attempt_start
                    elements_count = len(elements)
                    print(f"  🔍 Selector #{idx}: menemukan {elements_count} elemen (waktu: {selector_attempt_elapsed:.3f}s)")
                    
                    for btn_idx, btn in enumerate(elements, 1):
                        # Cek timeout di dalam loop juga
                        elapsed = time_module.time() - tutup_start_time
                        if elapsed >= max_tutup_time:
                            break
                            
                        try:
                            # Ambil text langsung dari button menggunakan JavaScript untuk menghindari nested text
                            # Gunakan textContent yang hanya mengambil text node langsung, bukan dari child elements
                            button_direct_text = driver.execute_script("""
                                var el = arguments[0];
                                // Ambil textContent yang hanya mengambil text langsung dari element ini
                                var directText = '';
                                if (el.childNodes) {
                                    for (var i = 0; i < el.childNodes.length; i++) {
                                        if (el.childNodes[i].nodeType === 3) { // Text node
                                            directText += el.childNodes[i].textContent || '';
                                        }
                                    }
                                }
                                // Jika tidak ada text node langsung, coba innerText sebagai fallback
                                if (!directText || directText.trim() === '') {
                                    directText = el.innerText || el.textContent || '';
                                }
                                return directText.trim();
                            """, btn) or ''
                            
                            # Fallback: gunakan .text jika JavaScript gagal
                            if not button_direct_text:
                                button_direct_text = btn.text or ''
                            
                            normalized_text = button_direct_text.strip().upper()
                            full_text = btn.text or ''  # Untuk debug saja
                            
                            # Cek visible dan enabled
                            if btn.is_displayed() and btn.is_enabled():
                                # Cek apakah text mengandung TUTUP atau exact match setelah normalize
                                if 'TUTUP' in normalized_text:
                                    # Untuk button dengan banyak nested text, cek apakah text langsung adalah "TUTUP" atau sangat pendek
                                    # Juga terima jika text langsung < 15 chars (lebih longgar untuk button di dalam modal dengan banyak text)
                                    # Atau jika text dimulai/berakhir dengan "TUTUP"
                                    text_length = len(normalized_text)
                                    is_exact_match = normalized_text == 'TUTUP'
                                    is_short_text = text_length < 15
                                    starts_with_tutup = normalized_text.startswith('TUTUP')
                                    ends_with_tutup = normalized_text.endswith('TUTUP')
                                    
                                    # Terima jika exact match, atau text pendek, atau text yang dimulai/diakhiri dengan TUTUP
                                    if is_exact_match or is_short_text or starts_with_tutup or ends_with_tutup:
                                        target_tutup = btn
                                        elapsed = time_module.time() - tutup_start_time
                                        print(f"  ✅ Element #{btn_idx} cocok!")
                                        print(f"     Text langsung: '{button_direct_text[:100]}...' (length: {len(button_direct_text)})")
                                        print(f"     Text normalized: '{normalized_text}'")
                                        print(f"     Match reason: exact={is_exact_match}, short={is_short_text}, starts={starts_with_tutup}, ends={ends_with_tutup}")
                                        print(f"✅ Tombol TUTUP ditemukan dengan selector #{idx} (waktu: {elapsed:.2f} detik)")
                                        break
                                    else:
                                        print(f"  ⚠️ Element #{btn_idx} mengandung TUTUP tapi tidak memenuhi kriteria: '{normalized_text[:50]}...' (length: {text_length}, skip)")
                                else:
                                    print(f"  ⚠️ Element #{btn_idx} tidak mengandung TUTUP: '{normalized_text[:50]}...'")
                            else:
                                print(f"  ⚠️ Element #{btn_idx} tidak displayed/enabled: '{normalized_text[:50]}...'")
                        except Exception as btn_e:
                            print(f"  ⚠️ Error saat cek element #{btn_idx}: {str(btn_e)[:50]}")
                            continue
                    
                    if target_tutup:
                        break
                except Exception as sel_e:
                    selector_attempt_elapsed = time_module.time() - selector_attempt_start
                    print(f"  ❌ Selector #{idx} error (waktu: {selector_attempt_elapsed:.3f}s): {str(sel_e)[:80]}")
                    continue
            
            # Fallback tambahan: Cari button di dalam modal/dialog yang mungkin berada di bagian bawah
            if not target_tutup:
                elapsed = time_module.time() - tutup_start_time
                if elapsed < max_tutup_time:
                    try:
                        print(f"  🔄 Mencoba mencari TUTUP di modal/dialog (fallback 2)...")
                        # Cari button di dalam modal, dialog, atau popup yang biasanya berada di bagian bawah
                        modal_selectors = [
                            (By.XPATH, "//div[contains(@class, 'modal') or contains(@class, 'dialog') or contains(@class, 'popup')]//button"),
                            (By.XPATH, "//div[@role='dialog']//button"),
                            (By.XPATH, "//div[contains(@class, 'mantine-Modal')]//button"),
                        ]
                        for modal_sel_type, modal_sel_value in modal_selectors:
                            try:
                                modal_buttons = driver.find_elements(modal_sel_type, modal_sel_value)
                                # Ambil button yang kemungkinan besar adalah TUTUP (biasanya yang terakhir atau di bagian bawah)
                                for modal_btn in reversed(modal_buttons):  # Mulai dari yang terakhir
                                    if modal_btn.is_displayed() and modal_btn.is_enabled():
                                        # Cek text langsung dengan JavaScript
                                        btn_direct_text = driver.execute_script("""
                                            var el = arguments[0];
                                            var directText = '';
                                            if (el.childNodes) {
                                                for (var i = 0; i < el.childNodes.length; i++) {
                                                    if (el.childNodes[i].nodeType === 3) {
                                                        directText += el.childNodes[i].textContent || '';
                                                    }
                                                }
                                            }
                                            if (!directText || directText.trim() === '') {
                                                directText = el.innerText || el.textContent || '';
                                            }
                                            return directText.trim().toUpperCase();
                                        """, modal_btn) or ''
                                        
                                        if 'TUTUP' in btn_direct_text and (btn_direct_text == 'TUTUP' or len(btn_direct_text) < 15 or btn_direct_text.startswith('TUTUP') or btn_direct_text.endswith('TUTUP')):
                                            target_tutup = modal_btn
                                            elapsed = time_module.time() - tutup_start_time
                                            print(f"  ✅ Tombol TUTUP ditemukan di modal (waktu: {elapsed:.2f} detik)")
                                            break
                                if target_tutup:
                                    break
                            except:
                                continue
                    except Exception as modal_e:
                        print(f"  ⚠️ Error saat mencari TUTUP di modal: {str(modal_e)[:50]}")
            
            # Fallback: Gunakan JavaScript untuk mencari dan klik langsung tombol TUTUP
            if not target_tutup:
                elapsed = time_module.time() - tutup_start_time
                if elapsed < max_tutup_time:
                    try:
                        print(f"  🔄 Mencoba JavaScript fallback untuk mencari TUTUP...")
                        # Gunakan JavaScript untuk mencari semua button dan cek text-nya, return info untuk dicari lagi
                        tutup_info = driver.execute_script("""
                            var buttons = document.querySelectorAll('button, a, [role="button"]');
                            var results = [];
                            for (var i = 0; i < buttons.length; i++) {
                                var btn = buttons[i];
                                
                                // Ambil text langsung dari button (hanya text node langsung, bukan nested)
                                var directText = '';
                                if (btn.childNodes) {
                                    for (var j = 0; j < btn.childNodes.length; j++) {
                                        if (btn.childNodes[j].nodeType === 3) { // Text node
                                            directText += btn.childNodes[j].textContent || '';
                                        }
                                    }
                                }
                                // Fallback ke innerText jika tidak ada text node langsung
                                if (!directText || directText.trim() === '') {
                                    directText = btn.innerText || btn.textContent || '';
                                }
                                
                                var text = directText.trim();
                                var textUpper = text.toUpperCase();
                                var textLength = text.length;
                                
                                // Cek visibility
                                var style = window.getComputedStyle(btn);
                                var isVisible = style.display !== 'none' && 
                                               style.visibility !== 'hidden' && 
                                               style.opacity !== '0' &&
                                               !btn.disabled;
                                
                                if (isVisible && textUpper.includes('TUTUP')) {
                                    // Prioritaskan exact match, text pendek, atau yang dimulai/diakhiri dengan TUTUP
                                    var isExactMatch = textUpper === 'TUTUP';
                                    var isShortText = textLength < 15;
                                    var startsWithTutup = textUpper.indexOf('TUTUP') === 0;
                                    var endsWithTutup = textUpper.lastIndexOf('TUTUP') === (textLength - 5);
                                    
                                    if (isExactMatch || isShortText || startsWithTutup || endsWithTutup) {
                                        // Coba ambil id, class, atau attributes untuk selector
                                        var id = btn.id || '';
                                        var classes = btn.className || '';
                                        var tagName = btn.tagName.toLowerCase();
                                        results.push({
                                            text: text,
                                            textUpper: textUpper,
                                            id: id,
                                            classes: classes,
                                            tagName: tagName,
                                            index: i
                                        });
                                    }
                                }
                            }
                            return results;
                        """)
                        
                        if tutup_info:
                            print(f"  🪟 JavaScript menemukan {len(tutup_info)} kandidat tombol TUTUP")
                            for info in tutup_info:
                                elapsed = time_module.time() - tutup_start_time
                                if elapsed >= max_tutup_time:
                                    break
                                try:
                                    # Coba cari element berdasarkan info yang didapat
                                    search_selectors = []
                                    if info['id']:
                                        search_selectors.append(f"//{info['tagName']}[@id='{info['id']}']")
                                    if info['classes']:
                                        # Ambil class pertama
                                        first_class = info['classes'].split()[0] if info['classes'] else None
                                        if first_class:
                                            search_selectors.append(f"//{info['tagName']}[contains(@class, '{first_class}')]")
                                    # Fallback: cari dengan text
                                    search_selectors.append(f"//{info['tagName']}[normalize-space(.)='{info['text'].strip()}']")
                                    
                                    for search_sel in search_selectors:
                                        try:
                                            found_btn = driver.find_element(By.XPATH, search_sel)
                                            if found_btn.is_displayed() and found_btn.is_enabled():
                                                target_tutup = found_btn
                                                elapsed = time_module.time() - tutup_start_time
                                                print(f"✅ Tombol TUTUP ditemukan dengan JavaScript fallback (waktu: {elapsed:.2f} detik): '{info['textUpper']}'")
                                                break
                                        except:
                                            continue
                                    
                                    if target_tutup:
                                        break
                                except:
                                    continue
                    except Exception as js_e:
                        print(f"  ⚠️ JavaScript fallback error: {str(js_e)[:80]}")
            
            if target_tutup:
                print("🚀 Mengklik tombol TUTUP...")
                clicked = False
                
                # Coba klik dengan beberapa metode (prioritaskan yang cepat)
                for attempt in range(2):  # Kurangi dari 3 ke 2 attempt untuk lebih cepat
                    try:
                        if attempt == 0:
                            # Coba normal click (paling cepat)
                            target_tutup.click()
                        else:
                            # Fallback: JavaScript click (jika normal click gagal)
                            driver.execute_script("arguments[0].click();", target_tutup)
                        
                        clicked = True
                        print(f"✅ Tombol TUTUP berhasil diklik (attempt {attempt + 1})")
                        break
                    except Exception as e:
                        if attempt == 0:
                            print(f"⚠️ Normal click gagal, mencoba JavaScript click...")
                        else:
                            print(f"❌ Gagal mengklik tombol TUTUP: {str(e)}")
                
                if clicked:
                    # Langsung klik ulang Catat Penjualan tanpa delay setelah klik TUTUP
                    print("🔄 Klik ulang 'Catat Penjualan' setelah TUTUP...")
                    try:
                        from .navigation_handler import click_catat_penjualan_direct
                        catat_reopen = click_catat_penjualan_direct(driver)
                        if catat_reopen:
                            print("✅ Berhasil klik ulang Catat Penjualan setelah TUTUP")
                        else:
                            print("⚠️ Gagal klik ulang Catat Penjualan, tapi akan lanjut")
                    except Exception as e_reopen:
                        print(f"⚠️ Error saat klik ulang Catat Penjualan: {str(e_reopen)}")
                    
                    print("✅ Popup TUTUP diklik dan Catat Penjualan sudah dibuka lagi - siap untuk NIK berikutnya")
                    return "REOPENED_AFTER_TUTUP"
                else:
                    print("❌ Gagal mengklik tombol TUTUP setelah semua attempt")
        except Exception as e:
            print(f"⚠️ Error saat mencari/mengklik tombol TUTUP: {str(e)}")
            pass

        # 3) Jika ada popup NIB dengan tombol "Nanti Saja, Lanjut Transaksi", klik tombol tersebut
        try:
            # Deteksi popup NIB dengan mencari teks "NIB" atau "Lengkapi NIB"
            nib_detected = False
            nib_selectors = [
                (By.XPATH, "//*[contains(translate(., 'abcdefghijklmnopqrstuvwxyz', 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'), 'LENGKAPI NIB')]"),
                (By.XPATH, "//*[contains(translate(., 'abcdefghijklmnopqrstuvwxyz', 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'), 'NOMOR INDUK BERUSAH')]"),
                (By.XPATH, "//*[contains(@class, 'icon-warning-nib')]"),
                (By.XPATH, "//*[contains(translate(., 'abcdefghijklmnopqrstuvwxyz', 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'), 'SEGERA LENGKAPI NIB')]")
            ]
            for how, sel in nib_selectors:
                if driver.find_elements(how, sel):
                    nib_detected = True
                    print("📋 Popup NIB terdeteksi — mencari tombol 'Nanti Saja, Lanjut Transaksi'...")
                    break
            
            if nib_detected:
                # Cari tombol "Nanti Saja, Lanjut Transaksi" atau "Lanjut Transaksi"
                lanjut_transaksi_buttons = driver.find_elements(By.XPATH, "//*[self::button or self::a][contains(translate(., 'abcdefghijklmnopqrstuvwxyz', 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'), 'LANJUT TRANSAKSI') or contains(translate(., 'abcdefghijklmnopqrstuvwxyz', 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'), 'NANTI SAJA')]")
                
                target_lanjut = None
                for btn in lanjut_transaksi_buttons:
                    label = (btn.text or '').strip().upper()
                    try:
                        if ('LANJUT' in label and 'TRANSAKSI' in label) or ('NANTI' in label and 'SAJA' in label):
                            if btn.is_displayed() and btn.is_enabled():
                                target_lanjut = btn
                                print(f"✅ Tombol 'Nanti Saja, Lanjut Transaksi' ditemukan: '{btn.text.strip()}'")
                                break
                    except:
                        continue
                
                if target_lanjut:
                    print("🚀 Mengklik tombol 'Nanti Saja, Lanjut Transaksi'...")
                    try:
                        target_lanjut.click()
                        print("✅ Tombol 'Nanti Saja, Lanjut Transaksi' berhasil diklik (normal click)")
                    except Exception as e:
                        try:
                            driver.execute_script("arguments[0].click();", target_lanjut)
                            print("✅ Tombol 'Nanti Saja, Lanjut Transaksi' berhasil diklik (JavaScript click)")
                        except Exception as e2:
                            print(f"❌ Gagal mengklik tombol 'Nanti Saja, Lanjut Transaksi': {str(e2)}")
                    
                    time.sleep(0.5)  # Tunggu popup tertutup
                    print("✅ Popup NIB ditutup - coba klik CEK PESANAN...")
                    # Setelah klik NIB, coba klik CEK PESANAN
                    try:
                        cek_pesanan_success = click_cek_pesanan(driver)
                        if cek_pesanan_success:
                            print("✅ CEK PESANAN berhasil diklik setelah handle NIB - lanjutkan proses")
                            return True
                        else:
                            print("⚠️ CEK PESANAN tidak ditemukan setelah handle NIB - lanjut cek popup lain")
                    except Exception as e:
                        print(f"⚠️ Error saat mencoba klik CEK PESANAN setelah handle NIB: {str(e)} - lanjut cek popup lain")
        except Exception as e:
            print(f"⚠️ Error saat mencari/mengklik popup NIB: {str(e)}")
            pass

        # 4) Jika tidak ada Ganti Pelanggan, tidak ada TUTUP, dan tidak ada NIB, coba klik CEK PESANAN lagi
        print("✅ Tidak ada popup/error terdeteksi - coba klik CEK PESANAN...")
        try:
            # Coba klik CEK PESANAN setelah handle popup (jika ada)
            cek_pesanan_success = click_cek_pesanan(driver)
            if cek_pesanan_success:
                print("✅ CEK PESANAN berhasil diklik setelah handle popup - lanjutkan proses")
                return True
            else:
                print("⚠️ CEK PESANAN masih tidak ditemukan setelah handle popup")
                return False
        except Exception as e:
            print(f"⚠️ Error saat mencoba klik CEK PESANAN setelah handle popup: {str(e)}")
            return False
        
    except Exception as e:
        print(f"❌ Error mengisi form NIK: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def click_cek_pesanan(driver):
    """
    ============================================
    KLIK TOMBOL "CEK PESANAN"
    ============================================
    Mencari dan mengklik tombol CEK PESANAN pada halaman Penjualan.
    """
    import time as time_module
    start_time = time_module.time()
    
    print("\n🧾 === KLIK CEK PESANAN ===")
    try:
        # Prioritas selector: yang paling cepat dulu
        candidate_selectors = [
            (By.XPATH, "//button[contains(translate(., 'abcdefghijklmnopqrstuvwxyz', 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'), 'CEK PESANAN')]", True),  # Text-based dengan wait dinamis (prioritas)
            (By.XPATH, "//html[1]/body[1]/div[1]/div[1]/div[1]/main[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/form[1]/div[4]/div[1]/button[1]", False),  # XPath langsung, no wait (fallback cepat)
        ]

        target = None
        max_total_time = 0.8  # Maksimal total waktu pencarian: 0.8 detik (dipercepat lagi)
        selector_index = 0
        for how, value, use_wait in candidate_selectors:
            selector_index += 1
            selector_start_time = time_module.time()
            
            # Cek waktu total, jika sudah melewati batas, langsung stop dan return
            elapsed_time = time_module.time() - start_time
            if elapsed_time >= max_total_time:
                print(f"⏱️ Waktu pencarian melewati batas {max_total_time} detik - stop mencari dan lanjut ke flow popup")
                elapsed_time = time_module.time() - start_time
                print(f"❌ Tombol CEK PESANAN tidak ditemukan (waktu: {elapsed_time:.2f} detik) - lanjut ke flow popup")
                return False  # Langsung return, skip debug info yang lambat
                
            try:
                selector_type = "WebDriverWait" if use_wait else "find_elements"
                print(f"🔍 Selector #{selector_index} ({selector_type}): Mencoba selector {how}...")
                
                # Coba dengan explicit wait (untuk selector berbasis text)
                if use_wait:
                    try:
                        # Hitung sisa waktu yang tersedia
                        remaining_time = max_total_time - (time_module.time() - start_time)
                        if remaining_time <= 0.05:
                            # Tidak ada waktu tersisa (kurang dari 0.05 detik), skip selector ini
                            print(f"  ⚠️ Skip selector #{selector_index}: tidak ada waktu tersisa (remaining: {remaining_time:.3f}s)")
                            continue
                        
                        # Gunakan wait yang sesuai dengan sisa waktu (maksimal 0.3 detik untuk cepat)
                        wait_timeout = min(0.3, max(0.05, remaining_time - 0.05))  # Kurangi 0.05 untuk margin
                        print(f"  ⏱️ WebDriverWait timeout: {wait_timeout:.3f} detik (remaining time: {remaining_time:.3f}s)")
                        
                        wait_remaining = WebDriverWait(driver, wait_timeout)
                        wait_result = None
                        try:
                            # Gunakan visibility_of_element_located untuk memastikan elemen benar-benar visible dan clickable
                            wait_remaining.until(EC.visibility_of_element_located((how, value)))
                            wait_result = "SUCCESS"
                            wait_elapsed = time_module.time() - selector_start_time
                            print(f"  ✅ WebDriverWait berhasil menemukan elemen (waktu: {wait_elapsed:.3f}s)")
                        except Exception as wait_e:
                            wait_result = f"TIMEOUT: {str(wait_e)[:50]}"
                            wait_elapsed = time_module.time() - selector_start_time
                            print(f"  ❌ WebDriverWait gagal (waktu: {wait_elapsed:.3f}s) - {wait_result}")
                            continue
                        
                        # Cari elemen setelah wait berhasil
                        elements = driver.find_elements(how, value)
                        elements_count = len(elements)
                        find_elapsed = time_module.time() - selector_start_time
                        print(f"  🔍 find_elements: menemukan {elements_count} elemen (waktu: {find_elapsed:.3f}s)")
                        
                        # Tunggu sebentar untuk memastikan elemen benar-benar siap (render selesai)
                        time_module.sleep(0.05)
                        
                        for idx, el in enumerate(elements):
                            try:
                                # Coba multiple method untuk mendapatkan text dengan retry
                                label = ""
                                for retry in range(3):  # Retry maksimal 3 kali
                                    try:
                                        # Coba el.text dulu
                                        label = (el.text or "").strip()
                                        if label:
                                            break
                                        # Jika kosong, coba get_attribute('textContent')
                                        label = (el.get_attribute('textContent') or "").strip()
                                        if label:
                                            break
                                        # Jika masih kosong, coba get_attribute('innerText')
                                        label = (el.get_attribute('innerText') or "").strip()
                                        if label:
                                            break
                                        # Jika masih kosong, tunggu sebentar dan coba lagi
                                        if retry < 2:
                                            time_module.sleep(0.05)
                                    except:
                                        if retry < 2:
                                            time_module.sleep(0.05)
                                
                                label_upper = label.upper()
                                # Validasi: 
                                # - Untuk selector berbasis text (selector 1): harus ada text yang valid
                                # - Untuk selector XPath langsung (selector 2): lebih fleksibel, bisa terima jika elemen ditemukan meskipun text belum ter-load
                                if selector_index == 1:
                                    # Selector pertama harus punya text yang valid
                                    is_valid = "CEK" in label_upper and "PESANAN" in label_upper
                                else:
                                    # Selector kedua (XPath langsung) lebih fleksibel - terima jika text valid atau label kosong
                                    is_valid = ("CEK" in label_upper and "PESANAN" in label_upper) or (not label)
                                
                                if is_valid:
                                    # Retry check is_displayed() karena elemen mungkin masih dalam proses render
                                    is_displayed = False
                                    for display_retry in range(5):  # Retry maksimal 5 kali
                                        try:
                                            is_displayed = el.is_displayed()
                                            if is_displayed:
                                                break
                                            # Jika belum displayed, tunggu sebentar
                                            if display_retry < 4:
                                                time_module.sleep(0.05)
                                        except:
                                            if display_retry < 4:
                                                time_module.sleep(0.05)
                                    
                                    if is_displayed:
                                        # Dapatkan posisi button
                                        try:
                                            location = el.location
                                            size = el.size
                                            position_info = f"x:{location['x']}, y:{location['y']}, width:{size['width']}, height:{size['height']}"
                                        except:
                                            position_info = "tidak dapat diambil"
                                        
                                        target = el
                                        selector_elapsed = time_module.time() - selector_start_time
                                        if label:
                                            print(f"  ✅ Element #{idx+1} cocok! Label: '{label}'")
                                        else:
                                            print(f"  ✅ Element #{idx+1} cocok! (label kosong, tapi selector spesifik dan elemen displayed)")
                                        print(f"  📍 Posisi button: {position_info}")
                                        print(f"  ✅ Tombol CEK PESANAN ditemukan dengan selector #{selector_index} ({selector_type}) - waktu: {selector_elapsed:.3f}s")
                                        break
                                    else:
                                        print(f"  ⚠️ Element #{idx+1} tidak displayed setelah retry: '{label if label else 'N/A'}'")
                                else:
                                    print(f"  ⚠️ Element #{idx+1} tidak cocok: label='{label[:30] if label else 'kosong'}...'")
                            except Exception as el_e:
                                print(f"  ⚠️ Error saat cek element #{idx+1}: {str(el_e)[:50]}")
                                continue
                        if target:
                            break
                    except Exception as wait_err:
                        # Jika explicit wait gagal, skip selector ini
                        wait_elapsed = time_module.time() - selector_start_time
                        print(f"  ❌ Selector #{selector_index} gagal (waktu: {wait_elapsed:.3f}s): {str(wait_err)[:80]}")
                        continue
                else:
                    # Langsung find_elements tanpa wait (untuk XPath langsung)
                    # Tapi tetap coba wait untuk visibility jika elemen ditemukan
                    elements = driver.find_elements(how, value)
                    elements_count = len(elements)
                    find_elapsed = time_module.time() - selector_start_time
                    print(f"  🔍 find_elements: menemukan {elements_count} elemen (waktu: {find_elapsed:.3f}s)")
                    
                    # Jika elemen ditemukan, coba wait untuk visibility dengan timeout singkat
                    if elements_count > 0:
                        try:
                            remaining_time = max_total_time - (time_module.time() - start_time)
                            if remaining_time > 0.1:  # Jika masih ada waktu tersisa
                                wait_timeout = min(0.2, remaining_time - 0.05)
                                wait_quick = WebDriverWait(driver, wait_timeout)
                                wait_quick.until(EC.visibility_of_element_located((how, value)))
                                time_module.sleep(0.05)  # Tunggu sebentar untuk memastikan render selesai
                        except:
                            pass  # Jika wait gagal, lanjutkan dengan elemen yang ada
                    
                    for idx, el in enumerate(elements):
                        try:
                            # Coba multiple method untuk mendapatkan text dengan retry
                            label = ""
                            for retry in range(3):  # Retry maksimal 3 kali
                                try:
                                    # Coba el.text dulu
                                    label = (el.text or "").strip()
                                    if label:
                                        break
                                    # Jika kosong, coba get_attribute('textContent')
                                    label = (el.get_attribute('textContent') or "").strip()
                                    if label:
                                        break
                                    # Jika masih kosong, coba get_attribute('innerText')
                                    label = (el.get_attribute('innerText') or "").strip()
                                    if label:
                                        break
                                    # Jika masih kosong, tunggu sebentar dan coba lagi
                                    if retry < 2:
                                        time_module.sleep(0.05)
                                except:
                                    if retry < 2:
                                        time_module.sleep(0.05)
                            
                            label_upper = label.upper()
                            # Validasi: 
                            # - Untuk selector berbasis text (selector 1): harus ada text yang valid
                            # - Untuk selector XPath langsung (selector 2): lebih fleksibel, bisa terima jika elemen ditemukan meskipun text belum ter-load
                            if selector_index == 1:
                                # Selector pertama harus punya text yang valid
                                is_valid = "CEK" in label_upper and "PESANAN" in label_upper
                            else:
                                # Selector kedua (XPath langsung) lebih fleksibel - terima jika text valid atau label kosong
                                is_valid = ("CEK" in label_upper and "PESANAN" in label_upper) or (not label)
                            
                            if is_valid:
                                # Retry check is_displayed() karena elemen mungkin masih dalam proses render
                                is_displayed = False
                                for display_retry in range(5):  # Retry maksimal 5 kali
                                    try:
                                        is_displayed = el.is_displayed()
                                        if is_displayed:
                                            break
                                        # Jika belum displayed, tunggu sebentar
                                        if display_retry < 4:
                                            time_module.sleep(0.05)
                                    except:
                                        if display_retry < 4:
                                            time_module.sleep(0.05)
                                
                                if is_displayed:
                                    # Dapatkan posisi button
                                    try:
                                        location = el.location
                                        size = el.size
                                        position_info = f"x:{location['x']}, y:{location['y']}, width:{size['width']}, height:{size['height']}"
                                    except:
                                        position_info = "tidak dapat diambil"
                                    
                                    target = el
                                    selector_elapsed = time_module.time() - selector_start_time
                                    if label:
                                        print(f"  ✅ Element #{idx+1} cocok! Label: '{label}'")
                                    else:
                                        print(f"  ✅ Element #{idx+1} cocok! (label kosong, tapi selector spesifik dan elemen displayed)")
                                    print(f"  📍 Posisi button: {position_info}")
                                    print(f"  ✅ Tombol CEK PESANAN ditemukan dengan selector #{selector_index} ({selector_type}) - waktu: {selector_elapsed:.3f}s")
                                    break
                                else:
                                    print(f"  ⚠️ Element #{idx+1} tidak displayed setelah retry: '{label if label else 'N/A'}'")
                            else:
                                print(f"  ⚠️ Element #{idx+1} tidak cocok: label='{label[:30] if label else 'kosong'}...'")
                        except Exception as el_e:
                            print(f"  ⚠️ Error saat cek element #{idx+1}: {str(el_e)[:50]}")
                            continue
                    if target:
                        break
            except Exception as e:
                selector_elapsed = time_module.time() - selector_start_time
                print(f"  ❌ Selector #{selector_index} error (waktu: {selector_elapsed:.3f}s): {str(e)[:80]}")
                continue

        elapsed_time = time_module.time() - start_time
        
        if not target:
            print(f"❌ Tombol CEK PESANAN tidak ditemukan setelah mencoba {selector_index} selector (waktu: {elapsed_time:.2f} detik) - lanjut ke flow popup")
            return False

        # Debug: Ambil informasi posisi dan detail button
        try:
            location = target.location
            size = target.size
            rect = target.rect
            print(f"📍 POSISI BUTTON CEK PESANAN:")
            print(f"   - Koordinat (x, y): ({location['x']}, {location['y']})")
            print(f"   - Ukuran (width, height): ({size['width']}, {size['height']})")
            print(f"   - Rect: {rect}")
            
            # Ambil XPath dan CSS untuk debugging
            xpath = driver.execute_script("""
                function absoluteXPath(el){
                  if(el.id) return '//*[@id="'+el.id+'"]';
                  const parts=[]; while(el && el.nodeType===1){
                    let ix=0, sib=el.previousSibling; while(sib){ if(sib.nodeType===1 && sib.nodeName===el.nodeName) ix++; sib=sib.previousSibling; }
                    parts.unshift(el.nodeName.toLowerCase()+'['+(ix+1)+']'); el=el.parentNode; }
                  return '//'+parts.join('/'); }
                return absoluteXPath(arguments[0]);
            """, target)
            css = driver.execute_script("""
                function cssPath(el){ if (!(el instanceof Element)) return; const path=[]; while (el.nodeType===1){ let selector=el.nodeName.toLowerCase(); if (el.id){ selector+='#'+el.id; path.unshift(selector); break; } else { let sib=el, nth=1; while (sib=sib.previousElementSibling){ if (sib.nodeName.toLowerCase()==selector) nth++; } selector += ':nth-of-type('+nth+')'; path.unshift(selector); el=el.parentNode; } } return path.join(' > '); }
                return cssPath(arguments[0]);
            """, target)
            print(f"🔗 CEK PESANAN XPath: {xpath}")
            print(f"🔗 CEK PESANAN CSS: {css}")
        except Exception as debug_err:
            print(f"⚠️ Error saat ambil debug info button: {str(debug_err)[:80]}")
        
        # Pastikan tombol visible dan dapat diklik
        try:
            # Scroll ke tombol jika perlu (tanpa delay)
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", target)
        except:
            pass
        
        print(f"✅ Tombol ditemukan: '{target.text.strip()}' — klik...")
        print(f"🔍 Info tombol: Displayed={target.is_displayed()}, Enabled={target.is_enabled()}")
        
        # Coba beberapa metode klik dengan tracking detail
        clicked = False
        click_method = None
        click_start_time = time_module.time()
        
        try:
            # Coba normal click dulu
            click_attempt_start = time_module.time()
            target.click()
            click_attempt_elapsed = time_module.time() - click_attempt_start
            clicked = True
            click_method = "normal click"
            print(f"✅ Tombol CEK PESANAN berhasil diklik (normal click, waktu: {click_attempt_elapsed:.3f}s)")
        except Exception as e1:
            normal_click_elapsed = time_module.time() - click_attempt_start
            print(f"  ❌ Normal click gagal (waktu: {normal_click_elapsed:.3f}s): {str(e1)[:80]}")
            try:
                # Fallback: JavaScript click
                js_click_start = time_module.time()
                driver.execute_script("arguments[0].click();", target)
                js_click_elapsed = time_module.time() - js_click_start
                clicked = True
                click_method = "JavaScript click"
                print(f"✅ Tombol CEK PESANAN berhasil diklik (JavaScript click, waktu: {js_click_elapsed:.3f}s)")
            except Exception as e2:
                js_click_elapsed = time_module.time() - js_click_start
                print(f"❌ Gagal klik tombol CEK PESANAN:")
                print(f"   - Normal click: {str(e1)[:80]} (waktu: {normal_click_elapsed:.3f}s)")
                print(f"   - JavaScript click: {str(e2)[:80]} (waktu: {js_click_elapsed:.3f}s)")
                return False
        
        if clicked:
            click_total_time = time_module.time() - click_start_time
            time.sleep(0.2)  # Kurangi delay dari 0.5 ke 0.2 detik
            elapsed_time = time_module.time() - start_time
            print(f"✅ Berhasil klik CEK PESANAN:")
            print(f"   - Metode klik: {click_method}")
            print(f"   - Waktu klik: {click_total_time:.3f}s")
            print(f"   - Total waktu pencarian + klik: {elapsed_time:.2f}s")
            return True
        else:
            elapsed_time = time_module.time() - start_time
            print(f"❌ Gagal klik CEK PESANAN (total waktu: {elapsed_time:.2f} detik)")
            return False

    except Exception as e:
        elapsed_time = time_module.time() - start_time
        print(f"❌ Error klik CEK PESANAN: {str(e)} (total waktu: {elapsed_time:.2f} detik)")
        return False


def click_proses_penjualan(driver):
    """
    ============================================
    KLIK TOMBOL "PROSES PENJUALAN" / "PROSES PESANAN"
    ============================================
    Pada halaman konfirmasi (Cek Penjualan), klik tombol proses.
    """
    print("\n🧾 === KLIK PROSES PENJUALAN ===")
    try:
        time.sleep(0.5)

        candidate_selectors = [
            (By.XPATH, "//html[1]/body[1]/div[1]/div[1]/div[1]/main[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[2]/div[1]/div[4]/div[1]/button[1]"),
            (By.XPATH, "//button[contains(translate(., 'abcdefghijklmnopqrstuvwxyz', 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'), 'PROSES PENJUALAN')]") ,
            (By.XPATH, "//button[contains(translate(., 'abcdefghijklmnopqrstuvwxyz', 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'), 'PROSES PESANAN')]") ,
            (By.XPATH, "//*[self::button or self::a][contains(translate(., 'abcdefghijklmnopqrstuvwxyz', 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'), 'PROSES')]"),
            (By.CLASS_NAME, "mantine-Button-root")
        ]

        target = None
        for how, value in candidate_selectors:
            try:
                elements = driver.find_elements(how, value)
                for el in elements:
                    label = (el.text or "").strip().upper()
                    if "PROSES" in label and ("PENJUALAN" in label or "PESANAN" in label) and el.is_enabled() and el.is_displayed():
                        target = el
                        break
                if target:
                    break
            except Exception:
                continue

        if not target:
            print("❌ Tombol PROSES PENJUALAN tidak ditemukan")
            return False

        try:
            xpath = driver.execute_script("""
                function absoluteXPath(el){
                  if(el.id) return '//*[@id="'+el.id+'"]';
                  const parts=[]; while(el && el.nodeType===1){
                    let ix=0, sib=el.previousSibling; while(sib){ if(sib.nodeType===1 && sib.nodeName===el.nodeName) ix++; sib=sib.previousSibling; }
                    parts.unshift(el.nodeName.toLowerCase()+'['+(ix+1)+']'); el=el.parentNode; }
                  return '//'+parts.join('/'); }
                return absoluteXPath(arguments[0]);
            """, target)
            css = driver.execute_script("""
                function cssPath(el){ if (!(el instanceof Element)) return; const path=[]; while (el.nodeType===1){ let selector=el.nodeName.toLowerCase(); if (el.id){ selector+='#'+el.id; path.unshift(selector); break; } else { let sib=el, nth=1; while (sib=sib.previousElementSibling){ if (sib.nodeName.toLowerCase()==selector) nth++; } selector += ':nth-of-type('+nth+')'; path.unshift(selector); el=el.parentNode; } } return path.join(' > '); }
                return cssPath(arguments[0]);
            """, target)
            print(f"🔗 PROSES PENJUALAN XPath: {xpath}")
            print(f"🔗 PROSES PENJUALAN CSS: {css}")
        except Exception:
            pass
        print(f"✅ Tombol ditemukan: '{target.text.strip()}' — klik...")
        target.click()
        time.sleep(0.5)
        print("✅ Berhasil klik PROSES PENJUALAN")
        return True

    except Exception as e:
        print(f"❌ Error klik PROSES PENJUALAN: {str(e)}")
        return False


def wait_for_captcha_and_success(driver, max_wait_seconds=180):
    """
    Menunggu user menyelesaikan captcha (slider) lalu mendeteksi halaman sukses transaksi.
    Strategi:
    - Tampilkan instruksi di terminal agar user menyelesaikan captcha.
    - Polling DOM tiap 0.5s untuk indikator sukses seperti 'LUNAS' atau tombol 'KIRIM STRUK KE PELANGGAN'.
    - Timeout default 180 detik.
    """
    print("\n🧩 === MENUNGGU USER SELESAIKAN CAPTCHA ===")
    try:
        print("Silakan selesaikan captcha pada popup (geser untuk cocokan gambar).")
        print("Setelah selesai, sistem akan otomatis mendeteksi halaman sukses.")

        waited = 0.0
        success_detected = False
        while waited < max_wait_seconds:
            try:
                # Indikator sukses potensial
                indicators = [
                    (By.XPATH, "//*[contains(translate(., 'abcdefghijklmnopqrstuvwxyz', 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'), 'LUNAS')]") ,
                    (By.XPATH, "//*[contains(translate(., 'abcdefghijklmnopqrstuvwxyz', 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'), 'KIRIM STRUK KE PELANGGAN')]") ,
                    (By.XPATH, "//*[contains(translate(., 'abcdefghijklmnopqrstuvwxyz', 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'), 'INFORMASI TRANSAKSI PELANGGAN')]") ,
                ]
                for how, sel in indicators:
                    els = driver.find_elements(how, sel)
                    if els:
                        success_detected = True
                        break
                if success_detected:
                    print("✅ Halaman sukses transaksi terdeteksi (LUNAS / Informasi Transaksi).")
                    return True
            except Exception:
                pass

            time.sleep(0.5)
            waited += 0.5

        print("❌ Timeout menunggu penyelesaian captcha / halaman sukses tidak terdeteksi.")
        return False

    except Exception as e:
        print(f"❌ Error saat menunggu captcha/sukses: {str(e)}")
        return False


def click_kembali_ke_halaman_utama(driver):
    """
    Klik tombol 'KEMBALI KE HALAMAN UTAMA' pada halaman sukses transaksi.
    Menggunakan direct text match dan fallback class.
    """
    print("\n🏠 === KLIK KEMBALI KE HALAMAN UTAMA ===")
    try:
        time.sleep(0.5)

        candidates = [
            (By.XPATH, "//*[self::button or self::a][contains(translate(., 'abcdefghijklmnopqrstuvwxyz', 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'), 'KEMBALI KE HALAMAN UTAMA')]") ,
            (By.XPATH, "//button[contains(@class,'button') and contains(translate(., 'abcdefghijklmnopqrstuvwxyz', 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'), 'KEMBALI')]") ,
            (By.CLASS_NAME, "mantine-Button-root")
        ]

        target = None
        for how, sel in candidates:
            try:
                els = driver.find_elements(how, sel)
                for el in els:
                    txt = (el.text or '').strip().upper()
                    if 'KEMBALI' in txt and 'HALAMAN' in txt:
                        if el.is_displayed() and el.is_enabled():
                            target = el
                            break
                if target:
                    break
            except Exception:
                continue

        if not target:
            print("❌ Tombol 'KEMBALI KE HALAMAN UTAMA' tidak ditemukan")
            return False

        # Cetak selector bantu
        try:
            xpath = driver.execute_script("""
                function absoluteXPath(el){ if(el.id) return '//*[@id="'+el.id+'"]';
                  const parts=[]; while(el && el.nodeType===1){ let ix=0, sib=el.previousSibling; while(sib){ if(sib.nodeType===1 && sib.nodeName===el.nodeName) ix++; sib=sib.previousSibling; }
                  parts.unshift(el.nodeName.toLowerCase()+'['+(ix+1)+']'); el=el.parentNode; } return '//'+parts.join('/'); }
                return absoluteXPath(arguments[0]);
            """, target)
            css = driver.execute_script("""
                function cssPath(el){ if (!(el instanceof Element)) return; const path=[]; while (el.nodeType===1){ let selector=el.nodeName.toLowerCase(); if (el.id){ selector+='#'+el.id; path.unshift(selector); break; } else { let sib=el, nth=1; while (sib=sib.previousElementSibling){ if (sib.nodeName.toLowerCase()==selector) nth++; } selector += ':nth-of-type('+nth+')'; path.unshift(selector); el=el.parentNode; } } return path.join(' > '); }
                return cssPath(arguments[0]);
            """, target)
            print(f"🔗 KEMBALI XPath: {xpath}")
            print(f"🔗 KEMBALI CSS: {css}")
        except Exception:
            pass

        print("🚀 Klik 'KEMBALI KE HALAMAN UTAMA'...")
        target.click()
        time.sleep(0.5)
        print("✅ Berhasil kembali ke halaman utama")
        return True

    except Exception as e:
        print(f"❌ Error klik kembali ke halaman utama: {str(e)}")
        return False

