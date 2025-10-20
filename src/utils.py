"""
Utility functions dan helper functions untuk automation script
File ini berisi fungsi-fungsi bantuan untuk file handling, validasi, dan logging
"""
import os
import time
import re
import pandas as pd
import requests
import logging
import queue
import threading
from datetime import datetime
from collections import Counter
from logging.handlers import RotatingFileHandler

# ========== KONFIGURASI PATH DIREKTORI ==========
# Path-path direktori untuk menyimpan file-file penting
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # Root directory project
AKUN_DIR = os.path.join(BASE_DIR, 'akun')                              # Direktori file akun Excel
RESULTS_DIR = os.path.join(BASE_DIR, 'results')                        # Direktori hasil export Excel
LOGS_DIR = os.path.join(BASE_DIR, 'logs')                              # Direktori file log
LOG_FILE = os.path.join(LOGS_DIR, 'automation.log')                    # File log utama

# ========== KONFIGURASI PROXY ==========
# Setting proxy untuk koneksi internet (jika diperlukan)
PROXY_CONFIG = {
    'host': '43.159.20.117',
    'port': '12233',
    'username': 'user-EdiFOSPr-region-sg',
    'password': 'FOS5610Proxy908',
    'auth_username': 'EdiFOSPr',
    'auth_password': 'FOS5610Proxy908'
}

# ========== KONSTANTA BULAN INDONESIA ==========
# Array nama bulan dalam bahasa Indonesia untuk format tanggal
BULAN_ID = [
    '', 'JANUARI', 'FEBRUARI', 'MARET', 'APRIL', 'MEI', 'JUNI',
    'JULI', 'AGUSTUS', 'SEPTEMBER', 'OKTOBER', 'NOVEMBER', 'DESEMBER'
]

# Setup logger untuk tracking error dan informasi debugging
logger = logging.getLogger('automation')

def setup_logging():
    """
    Setup konfigurasi logging untuk system automation
    Membuat file log dengan rotating handler untuk mencegah file terlalu besar
    """
    os.makedirs(LOGS_DIR, exist_ok=True)  # Buat direktori logs jika belum ada
    
    logger.setLevel(logging.DEBUG)  # Set level logging ke DEBUG untuk detail maksimal
    
    # Setup rotating file handler - file log akan di-rotate setiap 2MB
    handler = RotatingFileHandler(LOG_FILE, maxBytes=2*1024*1024, backupCount=3, encoding='utf-8')
    formatter = logging.Formatter('%(asctime)s %(levelname)s %(message)s')
    handler.setFormatter(formatter)
    logger.addHandler(handler)

def check_bandwidth():
    """Cek bandwidth proxy di https://faeyzaofficialstore.com/bandwidth"""
    print("\n🌐 Mengecek bandwidth proxy...")
    url = "https://faeyzaofficialstore.com/bandwidth"
    proxy = {
        'http': f'http://{PROXY_CONFIG["username"]}:{PROXY_CONFIG["password"]}@{PROXY_CONFIG["host"]}:{PROXY_CONFIG["port"]}',
        'https': f'http://{PROXY_CONFIG["username"]}:{PROXY_CONFIG["password"]}@{PROXY_CONFIG["host"]}:{PROXY_CONFIG["port"]}',
    }
    try:
        resp = requests.get(url, proxies=proxy, timeout=15, 
                          auth=(PROXY_CONFIG["auth_username"], PROXY_CONFIG["auth_password"]))
        if resp.status_code == 200:
            print("✅ Bandwidth proxy terhubung!")
            print(resp.text[:500])
        else:
            print(f"❌ Gagal cek bandwidth, status: {resp.status_code}")
    except Exception as e:
        print(f"❌ Error cek bandwidth: {e}")

def is_valid_email(username):
    """Validasi email sederhana"""
    return re.match(r"[^@]+@[^@]+\.[^@]+", username) is not None

def is_valid_phone(username):
    """Validasi nomor HP: hanya angka, panjang 10-15 digit, boleh leading zero"""
    return username.isdigit() and 10 <= len(username) <= 15 and (username.startswith('08') or username.startswith('628'))

def is_valid_pin(pin):
    """Validasi PIN"""
    return pin.isdigit() and 4 <= len(pin) <= 8

def load_accounts_from_excel(filename):
    """
    Load data akun dari file Excel dengan validasi lengkap
    Membaca file Excel dan memvalidasi setiap akun sebelum digunakan
    
    Args:
        filename (str): Path ke file Excel yang berisi data akun
        
    Returns:
        list: List tuple berisi (nama, username, pin) untuk akun yang valid
        
    Raises:
        ValueError: Jika tidak ada akun valid ditemukan di file Excel
    """
    # Baca file Excel dengan tipe data yang tepat untuk setiap kolom
    df = pd.read_excel(filename, dtype={"Nama": str, "Username": str, "Password": str})
    valid_accounts = []
    
    # Loop setiap baris dalam Excel untuk validasi dan ekstraksi data
    for _, row in df.iterrows():
        nama = str(row['Nama']).strip()      # Nama pangkalan (strip whitespace)
        username = str(row['Username']).strip()  # Username bisa email atau nomor HP
        pin = str(row['Password']).strip()       # PIN untuk login
        
        # Validasi username - harus berupa email atau nomor HP yang valid
        if not (is_valid_email(username) or is_valid_phone(username)):
            logger.warning(f"Invalid username (bukan email/HP): {username}")
            continue
            
        # Validasi PIN - harus berupa angka dengan panjang 4-8 digit
        if not is_valid_pin(pin):
            logger.warning(f"Invalid PIN: {pin} for {username}")
            continue
            
        # Jika semua validasi passed, tambahkan ke list akun valid
        valid_accounts.append((nama, username, pin))
    
    # Pastikan ada minimal satu akun valid
    if not valid_accounts:
        raise ValueError("No valid accounts found in Excel file!")
    
    return valid_accounts

def print_account_stats(accounts):
    """Print account statistics"""
    usernames = [acc[1] for acc in accounts]
    total = len(usernames)
    unique = len(set(usernames))
    
    print(f"\nTotal akun: {total}")
    print(f"Akun unik: {unique}")
    
    if total != unique:
        dupe = [u for u, c in Counter(usernames).items() if c > 1]
        print(f"Username duplikat: {dupe}")
    else:
        print("Tidak ada username yang duplikat.")

def input_with_timeout(prompt, timeout):
    """Input with timeout functionality"""
    q = queue.Queue()
    
    def inner():
        try:
            q.put(input(prompt))
        except Exception:
            q.put('')
    
    t = threading.Thread(target=inner)
    t.daemon = True
    t.start()
    
    try:
        return q.get(timeout=timeout)
    except queue.Empty:
        return ''

def get_main_menu_input():
    """Menu utama untuk memilih fitur program"""
    while True:
        try:
            print("\n🎯 === MENU UTAMA SNAPFLUX ====")
            print("Silakan pilih fitur yang ingin digunakan:")
            print("1. Check Stok")
            print("2. Batalkan Inputan")
            print("(Otomatis pilih 1 jika tidak ada input dalam 15 detik)")
            
            menu_input = input_with_timeout("Pilihan Anda (1/2): ", 15).strip()
            
            if not menu_input:
                print("⏭️ User tidak input menu - akan menggunakan Check Stok (default)")
                return 1
            
            if menu_input not in ['1', '2']:
                print("❌ Pilihan tidak valid! Silakan masukkan 1 atau 2")
                continue
            
            choice = int(menu_input)
            
            if choice == 1:
                print("✅ Dipilih: Check Stok")
                return 1
            elif choice == 2:
                print("✅ Dipilih: Batalkan Inputan")
                print("🚧 Fitur ini akan segera tersedia!")
                return 2
                
        except Exception as e:
            print(f"❌ Error input menu: {str(e)}")
            continue

def get_date_input():
    """Minta input tanggal dari user dengan validasi"""
    while True:
        try:
            print("\n📅 === FILTER TANGGAL LAPORAN PENJUALAN ===")
            print("Masukkan tanggal yang ingin diambil datanya (format: DD/MM/YYYY)")
            print("Contoh: 19/07/2025")
            print("Atau tekan Enter tanpa input untuk LEWATI filter tanggal")
            print("(Otomatis lewati filter tanggal jika tidak ada input dalam 15 detik)")
            
            date_input = input_with_timeout("Tanggal: ", 15).strip()
            
            if not date_input:
                print("⏭️ User tidak input tanggal - akan lewati fungsi klik tanggal")
                print("📅 Data akan diambil tanpa filter tanggal spesifik")
                return None
            
            if not re.match(r'^\d{2}/\d{2}/\d{4}$', date_input):
                print("❌ Format tanggal salah! Gunakan format DD/MM/YYYY")
                continue
            
            day, month, year = map(int, date_input.split('/'))
            
            if year < 2020 or year > 2030:
                print("❌ Tahun tidak valid! Gunakan tahun antara 2020-2030")
                continue
            if month < 1 or month > 12:
                print("❌ Bulan tidak valid!")
                continue
            if day < 1 or day > 31:
                print("❌ Hari tidak valid!")
                continue
            
            try:
                selected_date = datetime(year, month, day)
            except Exception:
                print("❌ Tanggal tidak valid!")
                continue
            
            print(f"✅ Menggunakan tanggal: {selected_date.strftime('%d %B %Y')}")
            return selected_date
            
        except Exception as e:
            print(f"❌ Error input tanggal: {str(e)}")

def get_excel_filename(selected_date=None):
    """Generate nama file Excel sesuai format"""
    if selected_date:
        tgl = selected_date.day
        bln = BULAN_ID[selected_date.month]
        thn = selected_date.year
        return f"DATA TRANSAKSI SNAPFLUX PANGKALAN {tgl} {bln} {thn}.xlsx"
    else:
        now = datetime.now()
        return f"DATA TRANSAKSI SNAPFLUX PANGKALAN TANPA FILTER TANGGAL {now.strftime('%d %B %Y')}.xlsx"

def parse_stok_to_int(stok_str):
    """Convert stok to integer"""
    if stok_str in ["Tidak Ditemukan", None, ""]:
        return None
    try:
        return int(re.findall(r'\d+', str(stok_str))[0])
    except:
        return None

def parse_inputan_to_int(inputan_str):
    """Convert inputan to integer (extract from '28 Tabung')"""
    if inputan_str in ["Tidak Ditemukan", None, ""]:
        return None
    try:
        return int(re.findall(r'\d+', str(inputan_str))[0])
    except:
        return None

def save_to_excel_pivot_format(pangkalan_id, nama_pangkalan, tanggal_check, stok_awal, total_inputan, status, selected_date=None):
    """Simpan data ke Excel dengan format pivot yang diminta user"""
    
    # File untuk pivot format
    filename = "DATA_TRANSAKSI_SNAPFLUX_HISTORIS_PIVOT.xlsx"
    filepath = os.path.join(RESULTS_DIR, filename)
    
    try:
        # Parse data
        stok_int = parse_stok_to_int(stok_awal)
        inputan_int = parse_inputan_to_int(total_inputan)
        
        # Format tanggal untuk header kolom sesuai permintaan user (tanpa ----)
        if selected_date:
            date_header = selected_date.strftime("%Y-%m-%d")
            display_date = selected_date.strftime("%Y-%m-%d")
        else:
            date_header = datetime.now().strftime("%Y-%m-%d")
            display_date = datetime.now().strftime("%Y-%m-%d")
        
        # Timestamp untuk TIME
        timestamp = datetime.now().strftime("%H:%M")
        
        print(f"🔧 Parsing data: stok='{stok_awal}' -> {stok_int}, inputan='{total_inputan}' -> {inputan_int}")
        
        # Load existing Excel atau buat baru
        if os.path.exists(filepath):
            try:
                # Baca Excel yang sudah ada dengan openpyxl untuk format yang lebih baik
                from openpyxl import load_workbook
                wb = load_workbook(filepath)
                
                # Cek apakah sheet 'Pivot View' ada
                if 'Pivot View' not in wb.sheetnames:
                    ws = wb.create_sheet('Pivot View')
                else:
                    ws = wb['Pivot View']
                    
            except Exception as e:
                print(f"⚠️ Error membaca file, akan buat file baru: {e}")
                # Buat workbook baru
                from openpyxl import Workbook
                wb = Workbook()
                if 'Sheet' in wb.sheetnames:
                    wb.remove(wb['Sheet'])
                ws = wb.create_sheet('Pivot View')
        else:
            # Buat workbook baru
            from openpyxl import Workbook
            wb = Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            ws = wb.create_sheet('Pivot View')
        
        # Inisialisasi variabel
        date_exists = False
        date_col_start = 3  # Default untuk file baru
        data_start_row = 3  # Default untuk file baru
        
        # Cek apakah header sudah ada
        if ws.max_row == 0 or ws.cell(row=1, column=1).value is None:
            # Buat header baru sesuai format yang diminta
            # Row 1: PANGKALAN_ID | NAMA_PANGKALAN | ---- DATE ----
            ws.cell(row=1, column=1, value="PANGKALAN_ID")
            ws.cell(row=1, column=2, value="NAMA_PANGKALAN")
            ws.cell(row=1, column=3, value=display_date)
            
            # Row 2: (kosong) | (kosong) | STOK (TABUNG) | INPUT (TABUNG) | TIME
            ws.cell(row=2, column=1, value="")
            ws.cell(row=2, column=2, value="")
            ws.cell(row=2, column=3, value="STOK (TABUNG)")
            ws.cell(row=2, column=4, value="INPUT (TABUNG)")
            ws.cell(row=2, column=5, value="TIME")
            date_exists = True
        else:
            # Cek apakah tanggal ini sudah ada di header
            for col in range(1, ws.max_column + 1):
                header_cell = ws.cell(row=1, column=col).value
                if header_cell and display_date in str(header_cell):
                    date_exists = True
                    date_col_start = col
                    break
            
            if not date_exists:
                # Tambahkan kolom tanggal baru
                date_col_start = ws.max_column + 1
                ws.cell(row=1, column=date_col_start, value=display_date)
                ws.cell(row=2, column=date_col_start, value="STOK (TABUNG)")
                ws.cell(row=2, column=date_col_start + 1, value="INPUT (TABUNG)")
                ws.cell(row=2, column=date_col_start + 2, value="TIME")
                date_exists = True
            
            # Pastikan sub-header tersedia untuk tanggal yang sudah ada
            ws.cell(row=2, column=date_col_start, value="STOK (TABUNG)")
            ws.cell(row=2, column=date_col_start + 1, value="INPUT (TABUNG)")
            ws.cell(row=2, column=date_col_start + 2, value="TIME")
            
            # Cari row yang tepat untuk data ini
            for row in range(3, ws.max_row + 1):
                pangkal_id_cell = ws.cell(row=row, column=1).value
                if pangkal_id_cell == pangkalan_id:
                    data_start_row = row
                    break
                elif pangkal_id_cell is None or pangkal_id_cell == "":
                    data_start_row = row
                    break
            
            if data_start_row == 3:  # Jika tidak ditemukan row yang tepat
                data_start_row = ws.max_row + 1
        
        # Isi data sesuai format yang diminta
        # Row 3+: pangkalan_id | nama_pangkalan | stok | input | time
        pangkalan_exists = False
        
        for row in range(3, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == pangkalan_id:
                # Update existing row
                ws.cell(row=row, column=date_col_start, value=stok_int)
                ws.cell(row=row, column=date_col_start + 1, value=inputan_int)
                ws.cell(row=row, column=date_col_start + 2, value=timestamp)
                pangkalan_exists = True
                break
        
        if not pangkalan_exists:
            # Tambahkan row baru
            new_row = data_start_row
            ws.cell(row=new_row, column=1, value=pangkalan_id)
            ws.cell(row=new_row, column=2, value=nama_pangkalan)
            ws.cell(row=new_row, column=date_col_start, value=stok_int)
            ws.cell(row=new_row, column=date_col_start + 1, value=inputan_int)
            ws.cell(row=new_row, column=date_col_start + 2, value=timestamp)
        
        # Import untuk styling (sebelum digunakan)
        from openpyxl.styles import Alignment, Border, Side, PatternFill
        
        # Apply conditional formatting untuk stok > 90 dan input = 0 (warna kuning)
        target_row = None
        if stok_int is not None and inputan_int is not None:
            if stok_int > 90 and inputan_int == 0:
                # Cari row yang tepat untuk conditional formatting
                if not pangkalan_exists:
                    # Row baru yang baru saja dibuat
                    target_row = new_row
                else:
                    # Cari row yang di-update
                    for row in range(3, ws.max_row + 1):
                        if ws.cell(row=row, column=1).value == pangkalan_id:
                            target_row = row
                            break
                
                # Apply warna kuning untuk kondisi stok > 90 dan input = 0
                if target_row:
                    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    ws.cell(row=target_row, column=date_col_start).fill = yellow_fill      # STOK
                    ws.cell(row=target_row, column=date_col_start + 1).fill = yellow_fill  # INPUT
                    print(f"🟡 Applied yellow highlight: STOK={stok_int} (>90) && INPUT={inputan_int} (=0)")
        
        # Format headers dengan merge cell dan center alignment dengan fill color berdasarkan hari
        from openpyxl.utils import get_column_letter
        
        # Hanya center alignment untuk headers, tanpa bold atau background color
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Fungsi untuk menentukan warna berdasarkan hari dalam seminggu
        def get_day_color(date_str):
            """Return fill color berdasarkan hari dalam seminggu"""
            try:
                # Parse tanggal dari format YYYY-MM-DD
                date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                weekday = date_obj.weekday()  # 0=Monday, 1=Tuesday, ..., 6=Sunday
                
                # Warna untuk setiap hari dalam seminggu
                day_colors = {
                    0: "FFE6B3",  # Senin - Jingga muda
                    1: "FFFFB3",  # Selasa - Kuning muda  
                    2: "FFB3B3",  # Rabu - Merah muda
                    3: "B3FFB3",  # Kamis - Hijau muda
                    4: "B3E6FF",  # Jumat - Biru muda
                    5: "E6B3FF",  # Sabtu - Ungu muda
                    6: "FFCCFF"   # Minggu - Pink muda
                }
                
                return day_colors.get(weekday, "FFFFFF")  # Default putih jika tidak ditemukan
            except:
                return "FFFFFF"  # Putih jika error parsing tanggal
        
        # Apply center alignment to row 1 dan 2 tanpa font bold atau background color
        for col in range(1, ws.max_column + 1):
            # Row 1 - hanya center alignment
            ws.cell(row=1, column=col).alignment = center_alignment
            # Row 2 - hanya center alignment  
            ws.cell(row=2, column=col).alignment = center_alignment
        
        # Merge cells untuk header tanggal (setiap grup tanggal memiliki 3 sub-kolom)
        # Temukan semua grup tanggal yang perlu di-merge
        date_groups = []
        
        # Cari semua kolom yang memiliki header tanggal (format YYYY-MM-DD)
        date_start_cols = []
        import re
        
        for col in range(1, ws.max_column + 1):
            header_cell = ws.cell(row=1, column=col).value
            if header_cell:
                # Deteksi apakah ini header tanggal dengan pattern YYYY-MM-DD
                is_date_header = re.match(r'^\d{4}-\d{2}-\d{2}$', str(header_cell))
                
                # Atau cek apakah ini bukan PANGKALAN_ID atau NAMA_PANGKALAN
                is_not_basic_header = str(header_cell) not in ["PANGKALAN_ID", "NAMA_PANGKALAN"]
                
                if is_date_header or (is_not_basic_header and col > 2):
                    # Cek apakah ini kolom pertama dari grup (harus diikuti STOK, INPUT, TIME di row 2)
                    if col + 2 <= ws.max_column:
                        row2_col1 = ws.cell(row=2, column=col).value
                        row2_col2 = ws.cell(row=2, column=col + 1).value
                        row2_col3 = ws.cell(row=2, column=col + 2).value
                        
                        if (row2_col1 == "STOK (TABUNG)" and 
                            row2_col2 == "INPUT (TABUNG)" and 
                            row2_col3 == "TIME"):
                            date_start_cols.append(col)
        
        # Buat grup merge untuk setiap tanggal
        for start_col in date_start_cols:
            end_col = min(start_col + 2, ws.max_column)
            date_groups.append((start_col, end_col))
        
        # Apply merge untuk setiap grup tanggal dengan fill color berdasarkan hari
        for start_col, end_col in date_groups:
            if end_col > start_col:
                try:
                    # Merge cell di row 1 untuk header tanggal
                    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                    
                    # Clear isi di kolom yang di-merge (kecuali kolom pertama)
                    for clear_col in range(start_col + 1, end_col + 1):
                        ws.cell(row=1, column=clear_col).value = None
                    
                    # Set alignment dan fill color untuk cell yang di-merge berdasarkan hari
                    merged_cell = ws.cell(row=1, column=start_col)
                    merged_cell.alignment = center_alignment
                    
                    # Dapatkan tanggal dari header untuk menentukan warna
                    header_value = merged_cell.value
                    if header_value and str(header_value).strip():
                        # Ambil tanggal dari header cell
                        date_str = str(header_value).strip()
                        
                        # Validasi apakah ini format tanggal YYYY-MM-DD
                        if re.match(r'^\d{4}-\d{2}-\d{2}$', date_str):
                            # Tentukan warna berdasarkan hari
                            fill_color = get_day_color(date_str)
                            
                            # Apply fill color
                            fill_pattern = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                            merged_cell.fill = fill_pattern
                            
                            # Juga apply fill color ke sub-header di row 2
                            for sub_col in range(start_col, end_col + 1):
                                sub_header_cell = ws.cell(row=2, column=sub_col)
                                sub_header_cell.fill = fill_pattern
                            
                            print(f"🎨 Applied color {fill_color} for date {date_str}")
                        
                except Exception as e:
                    print(f"⚠️ Warning: Failed to merge cells for date group {start_col}-{end_col}: {e}")
        
        # Pastikan semua header tanggal mendapat warna yang tepat (termasuk yang tidak di-merge)
        for col in range(1, ws.max_column + 1):
            header_cell = ws.cell(row=1, column=col)
            header_value = header_cell.value
            
            if header_value and str(header_value).strip():
                date_str = str(header_value).strip()
                
                # Validasi apakah ini format tanggal YYYY-MM-DD
                if re.match(r'^\d{4}-\d{2}-\d{2}$', date_str):
                    # Tentukan warna berdasarkan hari
                    fill_color = get_day_color(date_str)
                    fill_pattern = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    
                    # Apply fill color ke header tanggal
                    header_cell.fill = fill_pattern
                    
                    # Cari sub-header yang sesuai (3 kolom berikutnya: STOK, INPUT, TIME)
                    for sub_col_offset in range(3):
                        if col + sub_col_offset <= ws.max_column:
                            sub_header_cell = ws.cell(row=2, column=col + sub_col_offset)
                            sub_header_cell.fill = fill_pattern
                    
                    print(f"🎨 Applied color {fill_color} for date header {date_str} at column {col}")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Apply conditional formatting untuk semua data yang ada (stok > 90 dan input = 0)
        max_row = ws.max_row
        max_col = ws.max_column
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        print("🔍 Checking for conditional formatting conditions...")
        
        # Cari semua kolom yang berisi data stok dan input
        for col in range(1, max_col + 1):
            # Cek apakah ini kolom STOK (TABUNG) atau INPUT (TABUNG)
            row2_header = ws.cell(row=2, column=col).value
            if row2_header == "STOK (TABUNG)":
                # Ini kolom stok, cari kolom input yang bersebelahan (biasanya +1)
                input_col = col + 1
                if input_col <= max_col and ws.cell(row=2, column=input_col).value == "INPUT (TABUNG)":
                    # Loop melalui semua data rows (mulai dari row 3)
                    for row in range(3, max_row + 1):
                        try:
                            stok_value = ws.cell(row=row, column=col).value
                            input_value = ws.cell(row=row, column=input_col).value
                            
                            # Cek kondisi: stok > 90 dan input = 0
                            if stok_value is not None and input_value is not None:
                                stok_num = int(stok_value) if isinstance(stok_value, (int, str)) and str(stok_value).isdigit() else None
                                input_num = int(input_value) if isinstance(input_value, (int, str)) and str(input_value).isdigit() else None
                                
                                if stok_num is not None and input_num is not None:
                                    if stok_num > 90 and input_num == 0:
                                        # Apply warna kuning
                                        ws.cell(row=row, column=col).fill = yellow_fill      # STOK
                                        ws.cell(row=row, column=input_col).fill = yellow_fill # INPUT
                                        print(f"🟡 Applied yellow highlight at row {row}: STOK={stok_num} (>90) && INPUT={input_num} (=0)")
                        except (ValueError, TypeError):
                            # Skip jika tidak bisa convert ke int
                            continue
        
        # Apply border untuk semua cell
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Terapkan border ke semua cell yang memiliki data
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
        
        # Save workbook
        wb.save(filepath)
        
        print(f"✅ Data berhasil disimpan ke pivot format: {filepath}")
        print(f"📋 Format: PANGKALAN_ID={pangkalan_id}, NAMA={nama_pangkalan}")
        print(f"📋 Data: STOK={stok_int}, INPUT={inputan_int}, TIME={timestamp}")
        
    except Exception as e:
        print(f"❌ Error saat menyimpan pivot format: {str(e)}")
        logger.error(f"Error saving to pivot Excel: {str(e)}", exc_info=True)

def save_to_excel_new_format(nama_pangkalan, tanggal_check, stok_awal, total_inputan, status, selected_date=None, pangkalan_id=None):
    """Simpan data ke Excel dengan format baru (backward compatibility)"""
    filename = get_excel_filename(selected_date)
    filepath = os.path.join(RESULTS_DIR, filename)
    
    try:
        print(f"💾 Menyimpan data ke Excel ({filename})...")
        
        # Siapkan data dengan format baru
        data = {
            'NAMA PANGKALAN': [nama_pangkalan],
            'TANGGAL CHECK': [tanggal_check],
            'STOK AWAL': [stok_awal or "Tidak Ditemukan"],
            'TOTAL INPUTAN': [total_inputan or "Tidak Ditemukan"],
            'STATUS': [status]
        }
        if pangkalan_id:
            data['PANGKALAN_ID'] = [pangkalan_id]
        
        df_new = pd.DataFrame(data)
        
        if os.path.exists(filepath):
            df_existing = pd.read_excel(filepath)
            
            # Tambahkan data baru ke DataFrame yang sudah ada
            df_existing = pd.concat([df_existing, df_new], ignore_index=True)
            
            df_existing.to_excel(filepath, index=False)
            print(f"✅ Data berhasil disimpan ke: {filepath}")
            print("\n📋 Hasil scraping:")
            print(df_new.to_string(index=False))
        else:
            df_new.to_excel(filepath, index=False)
            print(f"✅ Data berhasil disimpan ke: {filepath}")
            print("\n📋 Hasil scraping:")
            print(df_new.to_string(index=False))
    except Exception as e:
        print(f"❌ Error saat menyimpan: {str(e)}")
        logger.error(f"Error saving to Excel: {str(e)}", exc_info=True)

def print_final_summary(rekap, accounts, akun_durations, total_start):
    """Print final summary of the scraping process"""
    total_end = time.time()
    total_duration = total_end - total_start
    rata2 = (sum(akun_durations) / len(akun_durations)) if akun_durations else 0
    
    print(f"\n✅ Semua akun selesai diproses!")
    print(f"⏱️ Total waktu proses semua akun: {total_duration:.2f} detik")
    print(f"⏱️ Rata-rata waktu proses per akun: {rata2:.2f} detik")
    
    # Rekap hasil akhir
    print("\n===== REKAPITULASI HASIL AKHIR =====")
    print(f"Sukses: {len(rekap['sukses'])} akun")
    if rekap['sukses']:
        for uname in rekap['sukses']:
            print(f"  ✅ {uname}")
    
    print(f"Gagal Login: {len(rekap['gagal_login'])} akun")
    if rekap['gagal_login']:
        for uname, reason in rekap['gagal_login']:
            print(f"  ❌ {uname}: {reason}")
    
    print(f"Gagal Navigasi/Data Penjualan: {len(rekap['gagal_navigasi'])} akun")
    if rekap['gagal_navigasi']:
        for uname, reason in rekap['gagal_navigasi']:
            print(f"  ⚠️ {uname}: {reason}")
    
    print(f"Gagal Ambil Waktu/Customer: {len(rekap['gagal_waktu'])} akun")
    if rekap['gagal_waktu']:
        for uname, reason in rekap['gagal_waktu']:
            print(f"  ⚠️ {uname}: {reason}")
    
    print(f"Gagal Ambil Stok: {len(rekap['gagal_stok'])} akun")
    if rekap['gagal_stok']:
        for uname, reason in rekap['gagal_stok']:
            print(f"  ⚠️ {uname}: {reason}")
    
    print(f"Error Lain: {len(rekap['error_lain'])} akun")
    if rekap['error_lain']:
        for uname, reason in rekap['error_lain']:
            print(f"  ❌ {uname}: {reason}")
    
    print("===== END REKAP =====\n")
    
    # Summary statistik
    total_accounts = len(accounts)
    success_rate = (len(rekap['sukses']) / total_accounts) * 100 if total_accounts > 0 else 0
    print(f"📊 STATISTIK FINAL:")
    print(f"Total Akun: {total_accounts}")
    print(f"Success Rate: {success_rate:.1f}%")
    print(f"Akun yang berhasil diproses: {len(rekap['sukses'])}")
    print(f"Akun yang gagal: {total_accounts - len(rekap['sukses'])}")