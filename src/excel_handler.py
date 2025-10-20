"""
Excel handler untuk mengelola operasi Excel
File ini berisi semua fungsi untuk menangani Excel operations
"""
import os
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from .constants import RESULTS_DIR, BULAN_ID, DAY_COLORS, EXCEL_FILENAME_PIVOT, SHEET_NAME_PIVOT
from .validators import parse_stok_to_int, parse_inputan_to_int

def get_excel_filename(selected_date=None):
    """Generate nama file Excel sesuai format"""
    if selected_date:
        tgl = selected_date.day
        bln = BULAN_ID[selected_date.month]
        thn = selected_date.year
        return f"DATA TRANSAKSI SNAPFLUX PANGKALAN {tgl} {bln} {thn}.xlsx"
    else:
        now = datetime.now()
        return f"DATA TRANSAKSI SNAPFLUX PANGKALAN TANPA FILTER TANGGAL {now.strftime('%d %B %Y')}.xlsx"

def get_day_color(date_str):
    """Return fill color berdasarkan hari dalam seminggu"""
    try:
        # Parse tanggal dari format YYYY-MM-DD
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        weekday = date_obj.weekday()  # 0=Monday, 1=Tuesday, ..., 6=Sunday
        return DAY_COLORS.get(weekday, "FFFFFF")  # Default putih jika tidak ditemukan
    except:
        return "FFFFFF"  # Putih jika error parsing tanggal

def save_to_excel_pivot_format(pangkalan_id, nama_pangkalan, tanggal_check, stok_awal, total_inputan, status, selected_date=None):
    """Simpan data ke Excel dengan format pivot yang diminta user"""
    
    # File untuk pivot format
    filename = EXCEL_FILENAME_PIVOT
    filepath = os.path.join(RESULTS_DIR, filename)
    
    try:
        # Parse data
        stok_int = parse_stok_to_int(stok_awal)
        inputan_int = parse_inputan_to_int(total_inputan)
        
        # Format tanggal untuk header kolom sesuai permintaan user (tanpa ----)
        if selected_date:
            date_header = selected_date.strftime("%Y-%m-%d")
            display_date = selected_date.strftime("%Y-%m-%d")
        else:
            date_header = datetime.now().strftime("%Y-%m-%d")
            display_date = datetime.now().strftime("%Y-%m-%d")
        
        # Timestamp untuk TIME
        timestamp = datetime.now().strftime("%H:%M")
        
        print(f"🔧 Parsing data: stok='{stok_awal}' -> {stok_int}, inputan='{total_inputan}' -> {inputan_int}")
        
        # Load existing Excel atau buat baru
        wb, ws = _load_or_create_workbook(filepath)
        
        # Inisialisasi variabel
        date_exists = False
        date_col_start = 3  # Default untuk file baru
        data_start_row = 3  # Default untuk file baru
        
        # Setup headers dan cek apakah tanggal sudah ada
        date_col_start, data_start_row = _setup_headers(ws, display_date, pangkalan_id)
        
        # Isi data sesuai format yang diminta
        _fill_data(ws, pangkalan_id, nama_pangkalan, stok_int, inputan_int, timestamp, date_col_start, data_start_row)
        
        # Apply conditional formatting untuk stok > 90 dan input = 0 (warna kuning)
        _apply_conditional_formatting(ws, stok_int, inputan_int, pangkalan_id, date_col_start)
        
        # Format headers dengan merge cell dan center alignment dengan fill color berdasarkan hari
        _format_headers(ws, display_date)
        
        # Apply border untuk semua cell
        _apply_borders(ws)
        
        # Save workbook
        wb.save(filepath)
        
        print(f"✅ Data berhasil disimpan ke pivot format: {filepath}")
        print(f"📋 Format: PANGKALAN_ID={pangkalan_id}, NAMA={nama_pangkalan}")
        print(f"📋 Data: STOK={stok_int}, INPUT={inputan_int}, TIME={timestamp}")
        
    except Exception as e:
        print(f"❌ Error saat menyimpan pivot format: {str(e)}")
        import logging
        logger = logging.getLogger('automation')
        logger.error(f"Error saving to pivot Excel: {str(e)}", exc_info=True)

def _load_or_create_workbook(filepath):
    """Load existing workbook atau buat yang baru"""
    if os.path.exists(filepath):
        try:
            wb = load_workbook(filepath)
            # Cek apakah sheet 'Pivot View' ada
            if SHEET_NAME_PIVOT not in wb.sheetnames:
                ws = wb.create_sheet(SHEET_NAME_PIVOT)
            else:
                ws = wb[SHEET_NAME_PIVOT]
                
        except Exception as e:
            print(f"⚠️ Error membaca file, akan buat file baru: {e}")
            # Buat workbook baru
            wb = Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            ws = wb.create_sheet(SHEET_NAME_PIVOT)
    else:
        # Buat workbook baru
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        ws = wb.create_sheet(SHEET_NAME_PIVOT)
    
    return wb, ws

def _setup_headers(ws, display_date, pangkalan_id):
    """Setup headers dan return column start dan row start"""
    date_col_start = 3  # Default untuk file baru
    data_start_row = 3  # Default untuk file baru
    
    # Cek apakah header sudah ada
    if ws.max_row == 0 or ws.cell(row=1, column=1).value is None:
        # Buat header baru sesuai format yang diminta
        # Row 1: PANGKALAN_ID | NAMA_PANGKALAN | ---- DATE ----
        ws.cell(row=1, column=1, value="PANGKALAN_ID")
        ws.cell(row=1, column=2, value="NAMA_PANGKALAN")
        ws.cell(row=1, column=3, value=display_date)
        
        # Row 2: (kosong) | (kosong) | STOK (TABUNG) | INPUT (TABUNG) | TIME
        ws.cell(row=2, column=1, value="")
        ws.cell(row=2, column=2, value="")
        ws.cell(row=2, column=3, value="STOK (TABUNG)")
        ws.cell(row=2, column=4, value="INPUT (TABUNG)")
        ws.cell(row=2, column=5, value="TIME")
        return date_col_start, data_start_row
    else:
        # Cek apakah tanggal ini sudah ada di header
        for col in range(1, ws.max_column + 1):
            header_cell = ws.cell(row=1, column=col).value
            if header_cell and display_date in str(header_cell):
                date_col_start = col
                break
        
        # Jika tanggal belum ada, tambahkan kolom baru
        if date_col_start == 3 and ws.cell(row=1, column=3).value != display_date:
            date_col_start = ws.max_column + 1
            ws.cell(row=1, column=date_col_start, value=display_date)
            ws.cell(row=2, column=date_col_start, value="STOK (TABUNG)")
            ws.cell(row=2, column=date_col_start + 1, value="INPUT (TABUNG)")
            ws.cell(row=2, column=date_col_start + 2, value="TIME")
        else:
            # Pastikan sub-header tersedia untuk tanggal yang sudah ada
            ws.cell(row=2, column=date_col_start, value="STOK (TABUNG)")
            ws.cell(row=2, column=date_col_start + 1, value="INPUT (TABUNG)")
            ws.cell(row=2, column=date_col_start + 2, value="TIME")
        
        # Cari row yang tepat untuk data ini
        for row in range(3, ws.max_row + 1):
            pangkal_id_cell = ws.cell(row=row, column=1).value
            if pangkal_id_cell == pangkalan_id:
                data_start_row = row
                break
            elif pangkal_id_cell is None or pangkal_id_cell == "":
                data_start_row = row
                break
        
        if data_start_row == 3 and ws.max_row > 2:  # Jika tidak ditemukan row yang tepat
            data_start_row = ws.max_row + 1
    
    return date_col_start, data_start_row

def _fill_data(ws, pangkalan_id, nama_pangkalan, stok_int, inputan_int, timestamp, date_col_start, data_start_row):
    """Fill data ke dalam worksheet"""
    pangkalan_exists = False
    
    for row in range(3, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == pangkalan_id:
            # Update existing row
            ws.cell(row=row, column=date_col_start, value=stok_int)
            ws.cell(row=row, column=date_col_start + 1, value=inputan_int)
            ws.cell(row=row, column=date_col_start + 2, value=timestamp)
            pangkalan_exists = True
            break
    
    if not pangkalan_exists:
        # Tambahkan row baru
        new_row = data_start_row
        ws.cell(row=new_row, column=1, value=pangkalan_id)
        ws.cell(row=new_row, column=2, value=nama_pangkalan)
        ws.cell(row=new_row, column=date_col_start, value=stok_int)
        ws.cell(row=new_row, column=date_col_start + 1, value=inputan_int)
        ws.cell(row=new_row, column=date_col_start + 2, value=timestamp)

def _apply_conditional_formatting(ws, stok_int, inputan_int, pangkalan_id, date_col_start):
    """Apply conditional formatting untuk stok > 90 dan input = 0"""
    if stok_int is not None and inputan_int is not None:
        if stok_int > 90 and inputan_int == 0:
            # Cari row yang tepat untuk conditional formatting
            for row in range(3, ws.max_row + 1):
                if ws.cell(row=row, column=1).value == pangkalan_id:
                    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    ws.cell(row=row, column=date_col_start).fill = yellow_fill      # STOK
                    ws.cell(row=row, column=date_col_start + 1).fill = yellow_fill  # INPUT
                    print(f"🟡 Applied yellow highlight: STOK={stok_int} (>90) && INPUT={inputan_int} (=0)")
                    break

def _format_headers(ws, display_date):
    """Format headers dengan merge cell dan center alignment dengan fill color berdasarkan hari"""
    # Hanya center alignment untuk headers, tanpa bold atau background color
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Apply center alignment to row 1 dan 2 tanpa font bold atau background color
    for col in range(1, ws.max_column + 1):
        # Row 1 - hanya center alignment
        ws.cell(row=1, column=col).alignment = center_alignment
        # Row 2 - hanya center alignment  
        ws.cell(row=2, column=col).alignment = center_alignment
    
    # Merge cells untuk header tanggal dan apply colors
    _merge_and_color_date_headers(ws, display_date, center_alignment)

def _merge_and_color_date_headers(ws, display_date, center_alignment):
    """Merge cells untuk header tanggal dan apply colors berdasarkan hari"""
    # Cari semua kolom yang memiliki header tanggal (format YYYY-MM-DD)
    date_start_cols = []
    
    for col in range(1, ws.max_column + 1):
        header_cell = ws.cell(row=1, column=col).value
        if header_cell:
            # Deteksi apakah ini header tanggal dengan pattern YYYY-MM-DD
            is_date_header = re.match(r'^\d{4}-\d{2}-\d{2}$', str(header_cell))
            is_not_basic_header = str(header_cell) not in ["PANGKALAN_ID", "NAMA_PANGKALAN"]
            
            if is_date_header or (is_not_basic_header and col > 2):
                # Cek apakah ini kolom pertama dari grup (harus diikuti STOK, INPUT, TIME di row 2)
                if col + 2 <= ws.max_column:
                    row2_col1 = ws.cell(row=2, column=col).value
                    row2_col2 = ws.cell(row=2, column=col + 1).value
                    row2_col3 = ws.cell(row=2, column=col + 2).value
                    
                    if (row2_col1 == "STOK (TABUNG)" and 
                        row2_col2 == "INPUT (TABUNG)" and 
                        row2_col3 == "TIME"):
                        date_start_cols.append(col)
    
    # Apply merge dan color untuk setiap grup tanggal
    for start_col in date_start_cols:
        end_col = min(start_col + 2, ws.max_column)
        if end_col > start_col:
            try:
                # Merge cell di row 1 untuk header tanggal
                ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                
                # Clear isi di kolom yang di-merge (kecuali kolom pertama)
                for clear_col in range(start_col + 1, end_col + 1):
                    ws.cell(row=1, column=clear_col).value = None
                
                # Set alignment dan fill color untuk cell yang di-merge berdasarkan hari
                merged_cell = ws.cell(row=1, column=start_col)
                merged_cell.alignment = center_alignment
                
                # Dapatkan tanggal dari header untuk menentukan warna
                header_value = merged_cell.value
                if header_value and str(header_value).strip():
                    date_str = str(header_value).strip()
                    
                    # Validasi apakah ini format tanggal YYYY-MM-DD
                    if re.match(r'^\d{4}-\d{2}-\d{2}$', date_str):
                        # Tentukan warna berdasarkan hari
                        fill_color = get_day_color(date_str)
                        
                        # Apply fill color
                        fill_pattern = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                        merged_cell.fill = fill_pattern
                        
                        # Juga apply fill color ke sub-header di row 2
                        for sub_col in range(start_col, end_col + 1):
                            sub_header_cell = ws.cell(row=2, column=sub_col)
                            sub_header_cell.fill = fill_pattern
                        
                        print(f"🎨 Applied color {fill_color} for date {date_str}")
                        
            except Exception as e:
                print(f"⚠️ Warning: Failed to merge cells for date group {start_col}-{end_col}: {e}")

def _apply_borders(ws):
    """Apply border untuk semua cell"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Terapkan border ke semua cell yang memiliki data
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
    
    # Apply conditional formatting untuk semua data yang ada
    _apply_all_conditional_formatting(ws)

def _apply_all_conditional_formatting(ws):
    """Apply conditional formatting untuk semua data yang ada"""
    max_row = ws.max_row
    max_col = ws.max_column
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    print("🔍 Checking for conditional formatting conditions...")
    
    # Cari semua kolom yang berisi data stok dan input
    for col in range(1, max_col + 1):
        # Cek apakah ini kolom STOK (TABUNG) atau INPUT (TABUNG)
        row2_header = ws.cell(row=2, column=col).value
        if row2_header == "STOK (TABUNG)":
            # Ini kolom stok, cari kolom input yang bersebelahan (biasanya +1)
            input_col = col + 1
            if input_col <= max_col and ws.cell(row=2, column=input_col).value == "INPUT (TABUNG)":
                # Loop melalui semua data rows (mulai dari row 3)
                for row in range(3, max_row + 1):
                    try:
                        stok_value = ws.cell(row=row, column=col).value
                        input_value = ws.cell(row=row, column=input_col).value
                        
                        # Cek kondisi: stok > 90 dan input = 0
                        if stok_value is not None and input_value is not None:
                            stok_num = int(stok_value) if isinstance(stok_value, (int, str)) and str(stok_value).isdigit() else None
                            input_num = int(input_value) if isinstance(input_value, (int, str)) and str(input_value).isdigit() else None
                            
                            if stok_num is not None and input_num is not None:
                                if stok_num > 90 and input_num == 0:
                                    # Apply warna kuning
                                    ws.cell(row=row, column=col).fill = yellow_fill      # STOK
                                    ws.cell(row=row, column=input_col).fill = yellow_fill # INPUT
                                    print(f"🟡 Applied yellow highlight at row {row}: STOK={stok_num} (>90) && INPUT={input_num} (=0)")
                    except (ValueError, TypeError):
                        # Skip jika tidak bisa convert ke int
                        continue

def save_to_excel_new_format(nama_pangkalan, tanggal_check, stok_awal, total_inputan, status, selected_date=None, pangkalan_id=None):
    """Simpan data ke Excel dengan format baru (backward compatibility)"""
    import pandas as pd
    
    filename = get_excel_filename(selected_date)
    filepath = os.path.join(RESULTS_DIR, filename)
    
    try:
        print(f"💾 Menyimpan data ke Excel ({filename})...")
        
        # Siapkan data dengan format baru
        data = {
            'NAMA PANGKALAN': [nama_pangkalan],
            'TANGGAL CHECK': [tanggal_check],
            'STOK AWAL': [stok_awal or "Tidak Ditemukan"],
            'TOTAL INPUTAN': [total_inputan or "Tidak Ditemukan"],
            'STATUS': [status]
        }
        if pangkalan_id:
            data['PANGKALAN_ID'] = [pangkalan_id]
        
        df_new = pd.DataFrame(data)
        
        if os.path.exists(filepath):
            df_existing = pd.read_excel(filepath)
            
            # Tambahkan data baru ke DataFrame yang sudah ada
            df_existing = pd.concat([df_existing, df_new], ignore_index=True)
            
            df_existing.to_excel(filepath, index=False)
            print(f"✅ Data berhasil disimpan ke: {filepath}")
            print("\n📋 Hasil scraping:")
            print(df_new.to_string(index=False))
        else:
            df_new.to_excel(filepath, index=False)
            print(f"✅ Data berhasil disimpan ke: {filepath}")
            print("\n📋 Hasil scraping:")
            print(df_new.to_string(index=False))
    except Exception as e:
        print(f"❌ Error saat menyimpan: {str(e)}")
        import logging
        logger = logging.getLogger('automation')
        logger.error(f"Error saving to Excel: {str(e)}", exc_info=True)
